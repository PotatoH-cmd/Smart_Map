"""FastAPI backend for Map Assistant"""
import os
import json
import logging
import asyncio
import math
import traceback
import re
import base64
import sqlite3
import uuid
import contextlib
import subprocess
import shutil
import tempfile
import threading
import time
import functools
from datetime import datetime as dt

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
except Exception:
    pass

# 在导入 torch 或 transformer 之前检查显卡设置
cuda_devices = os.environ.get("CUDA_VISIBLE_DEVICES", "Not Set")
logging.info(f"Current CUDA_VISIBLE_DEVICES: {cuda_devices}")
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, HTTPException, Header, Request, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, JSONResponse, Response, FileResponse, HTMLResponse
from starlette.types import Scope, Receive, Send
import httpx
from pydantic import BaseModel
from decimal import Decimal


class SelectiveGZipMiddleware(GZipMiddleware):
    """SSE 端点禁用 GZip，避免流式响应被缓冲后一次性返回。"""

    async def __call__(self, scope: Scope, receive: Receive, send: Send) -> None:
        if scope.get("type") == "http" and scope.get("path") == "/chat/stream":
            await self.app(scope, receive, send)
            return
        await super().__call__(scope, receive, send)
# Configure logging
LOG_FILE = "/home/server/python/map_assistant_v1/backend/backend.log"
# 确保文件存在且可写
with open(LOG_FILE, "a") as f:
    f.write(f"\n--- Service Restart at {dt.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n")

# 定义统一的格式
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# 创建文件处理器
file_handler = logging.FileHandler(LOG_FILE)
file_handler.setFormatter(formatter)

# 创建控制台处理器
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

# 配置根日志记录器
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)
root_logger.addHandler(file_handler)
root_logger.addHandler(console_handler)

logger = logging.getLogger(__name__)

# 特别确保 qwen_agent, tools 和 uvicorn 日志能写入文件
for name in ["qwen_agent", "qwen_agent_logger", "tools", "uvicorn", "uvicorn.access", "uvicorn.error", "httpx"]:
    l = logging.getLogger(name)
    if name in ["httpx", "qwen_agent", "qwen_agent_logger"]:
        l.setLevel(logging.DEBUG)  # 开启 DEBUG 级别以查看详细过程
    else:
        l.setLevel(logging.INFO)
    l.addHandler(file_handler)
    # l.propagate = False # 允许日志传播到控制台，方便调试工具调用
    l.setLevel(logging.INFO) if name not in ["httpx", "qwen_agent", "qwen_agent_logger"] else l.setLevel(logging.DEBUG)

# Import tools
from qwen_agent.agents import Assistant
from tools.map_tool import MapTool, LocationSearchTool
from tools.postgresql_tool import PostgreSQLTool
import psycopg2
# 知识库后端选择（环境变量 KNOWLEDGE_BACKEND: ragflow|llamaindex，默认 ragflow）
_kb_backend = os.environ.get("KNOWLEDGE_BACKEND", "ragflow")
if _kb_backend == "llamaindex":
    from tools.llamaindex_knowledge_tool import KnowledgeBaseTool
else:
    from tools.ragflow_knowledge_tool import KnowledgeBaseTool
from tools.knowledge_qa_agent import KnowledgeQAAgent
from tools.knowledge_graph_tool import get_kg  # 知识图谱工具
from tools.data_visualizer_tool import DataVisualizerTool
from tools.report_generator_tool import ReportGeneratorTool
from tools.weather_tool import WeatherTool
from tools.cesium_tool import CesiumTool  # Cesium 3D 地图工具
from tools.gis_tool_router import router as gis_tool_router  # GIS 处理工具
from tools.spatial_reference_tool import SpatialReferenceTool  # 空间参考工具（红线/采区），触发注册

# 切片管理服务包：注册表/统计缓存/构建/后台任务（原 main.py 中切片逻辑拆分）
from services.tile_manager import (
    _3DTILES_DATA_DIR,
    _3DTILES_REGISTRY_PATH,
    _CUSTOM_TILE_DATA_DIR,
    _DRONE_BUILD_JOBS,
    _DRONE_IMAGERY_DIR,
    _DRONE_MBTILES_DIR,
    _DRONE_REGISTRY_PATH,
    _DRONE_WORK_DIR,
    _MAX_3DTILES_FILES,
    _MAX_3DTILES_UNZIP_BYTES,
    _MAX_3DTILES_ZIP_BYTES,
    _OVERLAY_DATA_DIR,
    _TILE_BUILD_JOBS,
    _TILE_LAYER_META,
    _TILE_REGISTRY_PATH,
    _VT_DIR,
    _3dtiles_layer_to_row,
    _auto_register_existing_3dtiles_async,
    _copy_upload_limited,
    _count_3dtiles,
    _dir_stats,
    _drone_layer_to_row,
    _extract_zip_safely,
    _invalidate_tile_stats,
    _load_3dtiles_registry,
    _load_drone_registry,
    _load_tile_registry,
    _locate_tileset_root,
    _mbtiles_metadata,
    _media_type_for_tile_format,
    _merged_tile_meta,
    _parse_bounds,
    _parse_style,
    _read_3dtiles_meta,
    _register_drone_imagery,
    _run_drone_build_with_progress,
    _run_drone_mbtiles_build,
    _run_tippecanoe,
    _sanitize_layer_key,
    _save_3dtiles_registry,
    _save_drone_registry,
    _save_tile_registry,
    _submit_tile_build_job,
    _ttl_cache,
    _write_3dtiles_meta,
)

from agents import TaskExecutor
# 阶段1：RunEngine（可中断执行引擎）接入
from agents.run_engine import is_confirm_message, parse_supplied_from_message
from agents.run_store import get_run_store
# 阶段4：上下文预算（历史读库裁剪 / 滚动摘要压缩）
from agents.context_manager import load_history_from_db, COMPRESS_THRESHOLD_TURNS
# 阶段4：视图提示词 SSoT（2D/3D 唯一权威版本）
from prompts import build_view_system_message, build_elevation_injection
from cesium_bridge_server import cesium_ws_endpoint, get_cesium_client_count

from contextlib import asynccontextmanager

# ---------------------------------------------------------------------------
# SQLite 会话持久化
# ---------------------------------------------------------------------------
# 路径策略：优先环境变量 MAPASSIST_DB_PATH，默认项目 backend/sessions.db（不再硬编码旧机器路径）
DB_PATH = os.environ.get(
    "MAPASSIST_DB_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "sessions.db"),
)
_LEGACY_DB_PATH = "/home/server/python/map_assistant_v1/backend/sessions.db"
if not os.path.exists(DB_PATH) and os.path.exists(_LEGACY_DB_PATH):
    # 一次性迁移：旧机器硬编码路径存在而新路径不存在时复制过来
    try:
        shutil.copy2(_LEGACY_DB_PATH, DB_PATH)
        print(f"[init_db] 已从旧路径迁移 sessions.db: {_LEGACY_DB_PATH} -> {DB_PATH}")
    except Exception as e:  # noqa: BLE001
        print(f"[init_db] 旧 sessions.db 迁移失败（继续用新路径）: {e}")

# PostgreSQL 连接配置（环境变量注入，见 .env GEOSERVER_PG_*）
_PG_CONN = {
    "host": os.environ.get("GEOSERVER_PG_HOST", "172.136.16.52"),
    "port": int(os.environ.get("GEOSERVER_PG_PORT", "5432")),
    "dbname": os.environ.get("GEOSERVER_PG_DB", "postgres"),
    "user": os.environ.get("GEOSERVER_PG_USER", "postgres"),
    "password": os.environ.get("GEOSERVER_PG_PASSWORD", ""),
}

def init_db():
    with contextlib.closing(sqlite3.connect(DB_PATH)) as conn:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL DEFAULT '新对话',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL REFERENCES sessions(id) ON DELETE CASCADE,
                role TEXT NOT NULL,
                content TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_messages_session ON messages(session_id, id)"
        )
        # 阶段D：用户事实记忆（跨会话长期记忆，全局共享）
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_facts (
                id TEXT PRIMARY KEY,
                content TEXT NOT NULL,
                category TEXT,
                evidence TEXT,
                source_session TEXT,
                hits INTEGER DEFAULT 0,
                created_at TEXT,
                updated_at TEXT,
                last_seen_at TEXT
            )
        """)
        conn.commit()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    return conn

def now_iso():
    return dt.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


async def _daily_run_cleanup():
    """每日清理过期 run 数据（终态且超 7 天）。"""
    while True:
        await asyncio.sleep(86400)
        try:
            removed = get_run_store().cleanup(keep_days=7)
            if removed:
                logger.info(f"[run-cleanup] 已清理 {removed} 个过期 run 及其事件/检查点")
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[run-cleanup] 清理失败: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    global task_executor, bot
    init_db()
    # 启动时清理过期 run 数据，并启动每日定时清理
    try:
        removed = get_run_store().cleanup(keep_days=7)
        if removed:
            logger.info(f"[lifespan] 启动清理：移除 {removed} 个过期 run")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[lifespan] run 清理失败: {e}")
    app.state.run_cleanup_task = asyncio.create_task(_daily_run_cleanup())
    task_executor = init_task_executor()
    bot = init_agent()
    yield

app = FastAPI(title="Map Assistant API", version="1.0.0", lifespan=lifespan)

# Mount static files for reports
app.mount("/static", StaticFiles(directory="/home/server/python/map_assistant_v1/backend/static"), name="static")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(SelectiveGZipMiddleware, minimum_size=500)

# 注册 Cesium WebSocket 端点
app.add_api_websocket_route("/ws/cesium", cesium_ws_endpoint)

# 注册 GIS 处理工具路由
app.include_router(gis_tool_router)

bot = None
task_executor = None
LLM_CFG = None

class ChatMessage(BaseModel):
    role: str
    content: str
    images: Optional[List[str]] = None  # 图片 URL 列表（用于多模态消息）

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    active_view: str = 'map'  # 'map'(2D) | 'cesium'(3D) | 'kb'
    session_id: Optional[str] = None  # 可选会话 ID，用于 MemorySaver thread_id 隔离

class ScreenshotRequest(BaseModel):
    image_data: str
    file_name: Optional[str] = None

class KnowledgeItem(BaseModel):
    title: str
    content: str
    tags: List[str] = []

class KnowledgeQARequest(BaseModel):
    question: str
    top_k: int = 5

class KnowledgeAddRequest(BaseModel):
    name: str
    content: str

class ChatResponse(BaseModel):
    response: str
    messages: List[Dict[str, Any]]
    map_commands: List[Dict[str, Any]] = []
    cesium_commands: List[Dict[str, Any]] = []
    charts: List[Dict[str, Any]] = []
    report_url: Optional[str] = None
    intent_info: Optional[Dict[str, Any]] = None


def _extract_text_content(content) -> str:
    """从消息内容中提取纯文本。若为多模态数组，取第一个 text 部分。"""
    if isinstance(content, list):
        for part in content:
            if isinstance(part, dict) and part.get("type") == "text":
                return part.get("text", "")
        return ""
    return content or ""


# ============================================================
# 图片 OCR 预处理：使用 qwen-vl-ocr 提取文字后注入到对话
# ============================================================
OCR_MODEL = 'qwen-vl-ocr-2025-11-20'
OCR_BASE_URL = 'https://dashscope.aliyuncs.com/compatible-mode/v1'
OCR_API_KEY = 'sk-e4990da94bfb4037be1f755fa586d048'


UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "static")

async def _ocr_image(image_url: str) -> str:
    """使用 qwen-vl-ocr 专用模型从图片中提取文字内容。
    
    image_url 可以是：
    - 本地路径（如 /static/uploads/xxx.png），会自动转为 base64 data URL
    - 完整 HTTP(S) URL（DashScope 需要能访问）
    """
    # 将本地路径转为 base64 data URL（DashScope 无法访问内网地址）
    if image_url.startswith('/'):
        local_path = os.path.join(os.path.dirname(__file__), image_url.lstrip('/'))
        if not os.path.isfile(local_path):
            # 尝试相对于 static 目录
            local_path = os.path.join(UPLOAD_DIR, image_url.lstrip('/'))
        if os.path.isfile(local_path):
            with open(local_path, 'rb') as f:
                img_data = f.read()
            # 探测 MIME 类型
            ext = os.path.splitext(local_path)[1].lower()
            mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                        '.gif': 'image/gif', '.webp': 'image/webp', '.bmp': 'image/bmp'}
            mime = mime_map.get(ext, 'image/png')
            image_url = f"data:{mime};base64,{base64.b64encode(img_data).decode()}"
            logger.info(f"[OCR] 本地图片转 base64: {local_path} ({len(img_data)} bytes)")
        else:
            logger.error(f"[OCR] 图片文件不存在: {local_path}")
            return ""
    
    async with httpx.AsyncClient(timeout=120.0) as client:
        response = await client.post(
            f"{OCR_BASE_URL}/chat/completions",
            headers={
                "Authorization": f"Bearer {OCR_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": OCR_MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": "你是一个精确的文字提取工具。只输出图片中的纯文字内容，不要添加任何坐标框、边框标记、位置标注或额外说明。保持原始格式、精度和顺序。"
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": "提取图片中所有文字，只输出纯文本，不要任何坐标标注或边框信息。"
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": image_url},
                            }
                        ]
                    }
                ],
                "max_tokens": 4000,
            }
        )
        if response.status_code != 200:
            logger.error(f"[OCR] API 请求失败 HTTP {response.status_code}: {response.text[:300]}")
            return ""
        result = response.json()
        text = result["choices"][0]["message"]["content"]
        logger.info(f"[OCR] 提取成功 ({len(text)} 字符): {text[:200]}...")
        return text


async def _preprocess_excel(file_url: str) -> str:
    """使用 openpyxl 解析 Excel 文件，将表格数据格式化为文本。"""
    local_path = os.path.join(os.path.dirname(__file__), file_url.lstrip('/'))
    if not os.path.isfile(local_path):
        # 尝试相对于 static 目录
        local_path = os.path.join(UPLOAD_DIR, file_url.lstrip('/'))
    if not os.path.isfile(local_path):
        logger.error(f"[Excel] 文件不存在: {file_url} -> {local_path}")
        return "[Excel 文件读取失败]"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(local_path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row == 0:
                continue
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(cells))
            if rows:
                parts.append(f"【Excel 文件内容（{sheet_name}）】\n" + "\n".join(rows))
        wb.close()
        if parts:
            result = "\n\n".join(parts)
            logger.info(f"[Excel] 解析成功: {file_url} ({ws.max_row} 行)")
            return result
        return "[Excel 文件为空]"
    except Exception as e:
        logger.error(f"[Excel] 解析失败: {e}")
        return f"[Excel 文件解析失败: {e}]"


async def _preprocess_message_images(messages: list) -> list:
    """对用户消息中的附件进行预处理。
    
    - 图片：用 qwen-vl-ocr 提取文字
    - Excel：用 openpyxl 解析表格
    预处理结果注入到消息 content 中，并移除 images 字段。
    """
    for msg in messages:
        attachments = msg.get('images')
        if msg.get('role') != 'user' or not attachments or not isinstance(attachments, list):
            continue

        text_content = msg.get('content', '')
        all_results = []
        for file_url in attachments:
            is_excel = file_url.lower().endswith(('.xlsx', '.xls'))
            is_image = file_url.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp'))
            
            if is_excel:
                logger.info(f"[Excel] 开始处理: {file_url}")
                try:
                    result = await _preprocess_excel(file_url)
                    if result.strip():
                        all_results.append(result)
                except Exception as e:
                    logger.error(f"[Excel] 处理异常: {e}")
                    all_results.append(f"[Excel 读取失败: {e}]")
            elif is_image:
                logger.info(f"[OCR] 开始处理图片: {file_url}")
                try:
                    ocr_text = await _ocr_image(file_url)
                    if ocr_text.strip():
                        all_results.append(ocr_text)
                except Exception as e:
                    logger.error(f"[OCR] 图片处理异常: {e}")
                    all_results.append(f"[图片 OCR 失败: {e}]")
            else:
                logger.warning(f"[预处理] 跳过不支持的文件: {file_url}")

        if all_results:
            prefix = "【以下是从上传文件中自动提取的内容】\n"
            msg['content'] = text_content + "\n\n" + prefix + "\n---\n".join(all_results)
            logger.info(f"[预处理] 已将 {len(all_results)} 个文件内容注入到用户消息")
        
        # 移除 images 字段——主模型 qwen-flash 不支持，且文件内容已注入
        msg.pop('images', None)
    
    return messages


def init_task_executor():
    global LLM_CFG
    LLM_CFG = {
        'model': 'qwen-flash-2025-07-28',
        'model_server': 'https://dashscope.aliyuncs.com/compatible-mode/v1',
        'api_key': 'sk-e4990da94bfb4037be1f755fa586d048',
        'generate_cfg': {
            'extra_body': {
                'enable_thinking': False,
            },
        },
    }
    return TaskExecutor(LLM_CFG)


def init_agent():
    global LLM_CFG
    LLM_CFG = {
        'model': 'qwen-flash-2025-07-28',
        'model_server': 'https://dashscope.aliyuncs.com/compatible-mode/v1',
        'api_key': 'sk-e4990da94bfb4037be1f755fa586d048',
        'generate_cfg': {
            'extra_body': {
                'enable_thinking': False,
            },
        },
    }
    
    tools = [
        'map_tool',
        'location_search',
        'coordinate_marker',
        'postgresql_tool',
        'knowledge_base_tool',
        'data_visualizer_tool',
        'report_generator_tool',
        'weather_tool',
        'cesium_tool',  # Cesium 3D 地图工具
        'spatial_reference_tool',  # 空间参考数据工具（红线/采区边界等）
    ]
    
    return Assistant(
        llm=LLM_CFG,
        function_list=tools,
        name='Qwen3 地图助手',
        description="""你是一个专业的地图与数据分析助手，负责完成地图展示、数据库分析、政策查询和数据可视化任务。

------------------------------------------------------------
一、核心规则（最高优先级）
------------------------------------------------------------

1. 地图展示类任务
当用户提出以下需求时：
- "上图"
- "加载数据"
- "查看位置"
- "展示某砂场"

**根据当前视图选择工具（会话开头系统消息会指明）：**
- 2D 地图视图：调用 map_tool(action='load_vector_layer')
- 3D Cesium 视图：调用 cesium_tool(action='addGeoJsonLayer')
- 3D Cesium 视图且用户要求“测深风险/超深风险/风险柱/三维柱状展示”时：调用 cesium_tool(action='addDepthColumns')

如果是矢量图层加载，严格要求：
- 2D 模式必须使用 table_name='ceshen'
- 必须根据用户提到的名称设置 filter
- 严禁调用 location_search 或 add_marker
- 严禁在文本中返回经纬度

示例：

map_tool(
action='load_vector_layer',
table_name='ceshen',
filter="\"Mineable_Area_Name\"='种子场可采区'",
layer_name='种子场可采区'
)

严禁错误：
用户说 **种子场** → 加载 **潘庄砂场**

------------------------------------------------------------

2. 数据查询类任务

所有业务数据均来自 PostgreSQL 表：

ceshen

查询规则：

① 必须先调用

postgresql_tool(operation='get_db_schema')

② 字段必须加双引号，例如：

"Mineable_Area_Name"
"Measured_Depth"
"Control_Elevation"

③ 查询示例：

SELECT
AVG("Measured_Depth")
FROM ceshen

------------------------------------------------------------

3. 数据可视化任务

当用户提出：

- 生成图表
- 数据可视化
- 对比分析
- 统计并展示为柱状图/折线图/饼图

必须调用：

data_visualizer_tool

禁止行为：

- 禁止生成 Markdown 图片链接
- 禁止编写 Python 代码画图
- 禁止返回 Matplotlib 图

------------------------------------------------------------

4. 报告生成

只有用户明确提出：

- 生成报告
- 出具报告
- 形成文档

才允许调用：

report_generator_tool

报告生成流程（必须严格遵循）：

① 调用 knowledge_base_tool(operation='search', query='报告主题相关的政策、规范') 检索知识库
② 调用 postgresql_tool 获取业务数据（只使用真实查询结果，严禁虚构任何数据）
③ 调用 report_generator_tool 传入变量

【变量内容规范——必须严格遵守】

summary（摘要）：
- 简要说明报告背景、数据来源、整体结论
- 引用知识库中的相关政策或规范名称（不要复制原文，用自己的话概括）
- 只能使用数据库查询到的真实数据，不得捏造数字

details（详细数据）：
- 只填写真实的数据库查询结果
- 格式示例："查询返回 XX 条记录，其中 AA 字段均值为 BB，最大值为 CC..."
- 严禁虚构任何测量值、数量、地点或技术参数

knowledge_content（由工具自动填充）：
- 工具会自动从知识库检索并填充
- 无需手动传入，但如需传入，必须是经过归纳的文字，不得直接复制原始切片

conclusion（结论）：
- 结合知识库政策规范 + 数据库实际数据，得出综合结论
- 必须有依据，禁止主观推断或捏造

示例：
用户：生成固始县砂场超深度开采分析报告

正确流程：
1. knowledge_base_tool(operation='search', query='超深度开采判定规则 固始县')
2. postgresql_tool(operation='query', sql='SELECT "Mineable_Area_Name", AVG("Control_Elevation"-"Measured_Depth") as avg_depth FROM ceshen WHERE "County_District"=\'固始县\' GROUP BY "Mineable_Area_Name"')
3. report_generator_tool(variables={
     'report_title': '固始县砂场超深度开采分析报告',
     'summary': '本报告依据信阳市智慧巡河监管要求，对固始县XX个砂场进行超深度开采核查。（引用政策规范名称，不照搬原文）',
     'details': '数据库查询结果：固始县共有XX个砂场，平均超深XX米，最大超深砂场为XX（仅填真实查询值）',
     'conclusion': '综合以上数据与政策标准，建议...'
   })

------------------------------------------------------------

二、工具调用优先级（必须严格遵循）

工具调用顺序：

知识库 → 数据库 → 地图

规则：

① 政策 / 流程 / 操作文档问题

必须先调用

knowledge_base_tool(operation='search')

示例：

疏浚工程申请延期如何修改结束日期？

不得查询数据库。

② 数据统计问题

调用

postgresql_tool

③ 地图展示问题

根据当前视图选择工具：
- 2D 地图视图：调用 map_tool
- 3D Cesium 视图：调用 cesium_tool

如果会话开头有系统消息指明当前视图，必须严格按照指示选择工具。



------------------------------------------------------------

三、需求分析机制（必须执行）

在执行任务前必须进行 **需求分析**。

输出：

需求分析：
- 问题类型
- 目标输出
- 关键约束

示例：

需求分析

问题类型：数据分析  
目标输出：统计结果  
关键约束：采区名称

------------------------------------------------------------

四、任务拆解

根据需求分析生成执行计划：

子任务 | 工具 | 输出
---|---|---
查询数据库 | postgresql_tool | 数据表
生成图表 | data_visualizer_tool | 可视化图

------------------------------------------------------------

五、核心业务逻辑：实测高程 vs 控制高程

**警告：严禁将"实测高程"称为"实测深度"！两者物理意义相反！**

1. 字段物理意义：
   - Measured_Depth：代表"实测高程"（海拔高度）。
   - Control_Elevation：代表"控制高程"（准许挖掘的最低海拔）。

2. 判定逻辑：
   - 整体超深度开采定义：整个砂场的（实测高程-控制高程）测深点的 **平均差值超过 2米**，即定义为超深度开采。
     即：AVG(Control_Elevation - Measured_Depth) > 2。
   - 否则：视为整体合规（未构成超深度开采）。
   - 注意：如果平均差值未超过 2米，即使存在个别点位差值较大，也不能将该砂场定性为"超深度开采"。

3. 潘庄砂场专项修正：
   - 实测高程 (~30.5m) 远高于 控制高程 (17.1m)。
   - **结论：潘庄砂场完全不超深！** 它是未挖到控制深度，属于安全/合规状态。

4. 强制回答规范：
   - 若 平均差值 > 2m，回答格式："{区域}存在超深度开采。平均实测高程比控制高程低 {diff}m，超过 2m 允许范围。"
   - 否则，回答格式："{区域}未构成超深度开采。整体平均实测高程符合控制要求（平均偏差在 2m 以内）。"

------------------------------------------------------------

六、输出格式规范

回答必须结构化。

结构：

结论
概述
要点
建议

要求：

- 使用短句
- 使用 4–6 条要点
- 禁止长段落
- 先结论再解释

------------------------------------------------------------

七、禁止行为

以下行为绝对禁止：

1. 政策问题查询数据库
2. 地图加载调用 knowledge_base_tool
3. 未经请求主动上图
4. 生成 Python 图表
5. 返回 Markdown 图片

------------------------------------------------------------

八、交互确认

地图加载成功后，回复语必须与用户请求 **完全一致**。

示例：

"种子场可采区数据已成功加载到地图。"

------------------------------------------------------------

九、空间参考数据（红线/采区边界等）

系统内置了空间参考图层，可通过 spatial_reference_tool 查询：

| 图层 key | 名称 | 说明 |
|----------|------|------|
| hx | 河道管理红线 | 河道管理范围法定边界 |
| caiqu | 2025年可采区边界 | 许可采砂区域空间范围 |

**自动触发规则**：
- 用户提及"红线""河道红线""红线范围" → 自动关联 hx 图层
- 用户提及"采区""可采区""采砂区范围" → 自动关联 caiqu 图层

**使用流程**（当用户问题涉及空间参考数据时）：
1. 调用 spatial_reference_tool(action='get_geometry', layer='hx'或'caiqu') 获取参考几何
2. 将几何作为 WKT 传给 postgresql_tool(operation='spatial_query') 做空间筛选
3. 或传给 map_tool 将参考数据叠加到地图显示

示例：用户问"红线附近的采砂场有哪些"
步骤① spatial_reference_tool(action='get_geometry', layer='hx')
步骤② postgresql_tool(operation='spatial_query', spatial_table='ceshen', spatial_op='within', spatial_geom_wkt='<步骤①返回的几何WKT>')
        """
    )

# 工具实例缓存（延迟初始化，第一次调用时创建）
_TOOL_INSTANCES = {}

def _get_tool_instance(name):
    if name not in _TOOL_INSTANCES:
        if name == 'map_tool':
            _TOOL_INSTANCES[name] = MapTool()
        elif name == 'location_search':
            _TOOL_INSTANCES[name] = LocationSearchTool()
        elif name == 'postgresql_tool':
            _TOOL_INSTANCES[name] = PostgreSQLTool()
        elif name == 'knowledge_base_tool':
            _TOOL_INSTANCES[name] = KnowledgeBaseTool()
        elif name == 'data_visualizer_tool':
            _TOOL_INSTANCES[name] = DataVisualizerTool()
        elif name == 'report_generator_tool':
            _TOOL_INSTANCES[name] = ReportGeneratorTool()
        elif name == 'weather_tool':
            _TOOL_INSTANCES[name] = WeatherTool()
        elif name == 'cesium_tool':
            _TOOL_INSTANCES[name] = CesiumTool()
        elif name == 'spatial_reference_tool':
            from tools.spatial_reference_tool import SpatialReferenceTool
            _TOOL_INSTANCES[name] = SpatialReferenceTool()
        else:
            return name  # 未知工具回退字符串
    return _TOOL_INSTANCES[name]

def build_assistant_with_tools(function_list):
    # 优先使用工具实例，避免字符串名称查找失败
    resolved = [_get_tool_instance(name) for name in function_list]
    return Assistant(
        llm=LLM_CFG,
        function_list=resolved,
        name='Qwen3 地图助手',
        description="""地图数据加载任务，仅使用地图与数据库工具。"""
    )

@app.get("/")
async def root():
    return {"message": "Map Assistant API is running"}

# ==============================
# ✅ 优化：动态加载矢量数据接口
# ==============================
def _resolve_vector_filter(pg_tool, safe_table_name: str, filter_text: str,
                           table_cols_lower: dict) -> Optional[str]:
    """校验 filter 中引用的字段是否存在于目标表。

    - 字段是目标表真实列 → 保留原样；
    - 目标表是 jsonb 属性表（如 caiqu/hx）且 properties 中存在该键
      → 改写为 (properties->>'字段')；
    - 字段既不是列也不在 properties 中 → 丢弃整个 filter（加载全部要素），
      避免"字段不存在"导致整表查询失败。
    """
    refs = re.findall(r'"([A-Za-z_][A-Za-z0-9_]*)"', filter_text or "")
    resolved = filter_text
    for field in refs:
        if field.lower() in table_cols_lower:
            continue  # 真实列，保留原样
        if "properties" in table_cols_lower:
            try:
                probe = pg_tool.call({
                    "operation": "query",
                    "sql": f"SELECT properties ? %s AS has_key FROM {safe_table_name} LIMIT 1",
                    "params": [field],
                })
                row = (probe.get("data") or [{}])[0] if probe.get("success") else {}
                if row.get("has_key"):
                    resolved = resolved.replace(
                        f'"{field}"', f"(properties->>'{field}')"
                    )
                    continue
            except Exception as e:
                logger.warning(f"Vector API properties probe failed for '{safe_table_name}': {e}")
        logger.warning(
            f"Vector API dropping filter for table '{safe_table_name}': "
            f"field '{field}' not found in columns or properties"
        )
        return None
    return resolved


@app.get("/api/vector-data")
async def get_vector_data(
    table_name: str,
    geom_col: str = 'geom',
    properties: str = None,
    filter: str = None,
    color_expression: str = None,
    debug: bool = False
):
    """
    动态获取指定表的 GeoJSON 数据
    :param table_name: 数据库表名
    :param geom_col: 几何列名，默认为 'geom'
    :param properties: 需要包含在 properties 中的字段名，逗号分隔。
    :param filter: SQL 过滤条件 (WHERE 后的内容，如 "name='xxx'")
    :param color_expression: SQL 颜色表达式，例如 "CASE WHEN depth < 10 THEN 'red' ELSE 'blue' END"
    """
    # 兼容性处理：如果请求的是旧表名 mineable_areas，自动映射到新表 ceshen
    target_table = table_name.strip().lower()
    if target_table == 'mineable_areas' or target_table == '"mineable_areas"':
        logger.info(f"Redirecting table_name from '{table_name}' to 'ceshen'")
        table_name = 'ceshen'

    try:
        logger.info(f"Vector API request: table_name={table_name}, geom_col={geom_col}, properties={properties}, filter={filter}, color_expression={color_expression}, debug={debug}")
        # 安全性校验：允许字母、数字、下划线、双引号、单引号、等号、空格和中文字符
        # 注意：此处 filter 校验需要比较宽松，但也需防止恶意 SQL 注入
        if not re.match(r'^[a-zA-Z0-9_"\u4e00-\u9fa5\s\'\.\(\)\=\!\<\>\-\+]+$', table_name):
            raise HTTPException(status_code=400, detail="无效的表名格式")

        pg_tool = PostgreSQLTool(cfg={
            'host': '172.136.16.52',
            'port': 5432,
            'database': 'postgres',
            'user': 'postgres',
        })

        # 修复 color_expression 中的字段引用，增加表别名 t. 以避免字段不存在报错
        safe_color_expression = color_expression
        if color_expression:
            # 匹配双引号中的字段名，例如 "Measured_Depth" -> "t"."Measured_Depth"
            safe_color_expression = re.sub(r'("([a-zA-Z0-9_]+)")', r'"t".\1', color_expression)

        # 构建属性 JSON 对象
        if properties:
            props_list = [p.strip() for p in properties.split(',')]
            # 确保关键字段始终包含在内，用于前端 Popup 显示（仅限目标表实际存在的字段）
            essential_fields = [
                '"Mineable_Area_Name"', '"Measured_Depth"', '"Control_Elevation"',
                '"Lon_4326"', '"Lat_4326"', '"Year"', '"Mineable_Area_ID"', '"County_District"'
            ]
            for field in essential_fields:
                clean_field = field.replace('"', '')
                if clean_field.lower() in table_cols_lower and clean_field not in props_list:
                    # 用表中实际列名（兼容大小写）追加，避免引用不存在的字段导致查询失败
                    props_list.append(table_cols_lower[clean_field.lower()])
            
            # 修复：避免在 f-string 表达式中使用反斜杠
            formatted_props = []
            for p in props_list:
                if not p.startswith('"'):
                    formatted_props.append(f"'{p}', \"t\".\"{p}\"")
                else:
                    clean_p = p.replace('"', '')
                    formatted_props.append(f"'{clean_p}', \"t\".{p}")
            
            props_json = ", ".join(formatted_props)
            if safe_color_expression:
                props_json += f", '_style_color', {safe_color_expression}"
            props_sql = f"json_build_object({props_json})"
        else:
            if safe_color_expression:
                props_sql = f"(row_to_json(t)::jsonb - '{geom_col}' || jsonb_build_object('_style_color', {safe_color_expression}))::json"
            else:
                props_sql = f"(row_to_json(t)::jsonb - '{geom_col}')::json"

        safe_table_name = table_name if table_name.startswith('"') else f'"{table_name}"'

        # 查询目标表实际列名，动态决定是否启用经纬度回退（caiqu/hx 等 jsonb 表无 Lon_4326/Lat_4326 列）
        table_cols_lower = {}
        try:
            col_res = pg_tool.call({
                'operation': 'query',
                'sql': """
                    SELECT column_name FROM information_schema.columns
                    WHERE table_schema = 'public' AND LOWER(table_name) = LOWER(%s)
                """,
                'params': [table_name.strip('"')]
            })
            if col_res.get('success'):
                table_cols_lower = {str(r.get('column_name')).lower(): str(r.get('column_name'))
                                    for r in (col_res.get('data') or []) if r.get('column_name')}
        except Exception as e:
            logger.warning(f"Vector API failed to fetch columns for '{table_name}': {e}")
        has_lonlat = 'lon_4326' in table_cols_lower and 'lat_4326' in table_cols_lower
        lon_col = table_cols_lower.get('lon_4326') if has_lonlat else None
        lat_col = table_cols_lower.get('lat_4326') if has_lonlat else None
        lonlat_case = ""
        if has_lonlat:
            lonlat_case = (
                f'WHEN "{lon_col}" IS NOT NULL AND "{lat_col}" IS NOT NULL THEN\n'
                f'                                    ST_SetSRID(ST_MakePoint("{lon_col}", "{lat_col}"), 4326)\n'
            )

        # 处理过滤条件（包含几何或经纬度回退；经纬度回退仅对含经纬度列的表生效）
        where_geom_valid = f"({geom_col} IS NOT NULL)"
        if has_lonlat:
            where_lonlat_valid = f"(\"{lon_col}\" IS NOT NULL AND \"{lat_col}\" IS NOT NULL)"
            where_clause = f"WHERE ({where_geom_valid} OR {where_lonlat_valid})"
        else:
            where_clause = f"WHERE {where_geom_valid}"
        # 校验 filter 引用的字段是否存在于目标表；不存在时改写到 jsonb properties
        # 或直接丢弃 filter（加载全部要素），避免"字段不存在"导致整表查询失败。
        if filter:
            filter = _resolve_vector_filter(pg_tool, safe_table_name, filter, table_cols_lower)
        if filter:
            where_clause += f" AND ({filter})"

        def build_empty_meta(count_filter: str):
            count_sql = f"SELECT COUNT(*)::int AS cnt FROM {safe_table_name} AS t WHERE {count_filter};"
            geom_sql = f"SELECT COUNT(*)::int AS cnt FROM {safe_table_name} AS t WHERE ({count_filter}) AND ({geom_col} IS NOT NULL);"
            geom_valid_sql = f"SELECT COUNT(*)::int AS cnt FROM {safe_table_name} AS t WHERE ({count_filter}) AND ({geom_col} IS NOT NULL AND ST_IsValid({geom_col}));"
            lonlat_res = {'success': False, 'error': None}
            if has_lonlat:
                lonlat_sql = f"SELECT COUNT(*)::int AS cnt FROM {safe_table_name} AS t WHERE ({count_filter}) AND (\"{lon_col}\" IS NOT NULL AND \"{lat_col}\" IS NOT NULL);"
                lonlat_res = pg_tool.call({'operation': 'query', 'sql': lonlat_sql, 'params': []})
            count_res = pg_tool.call({'operation': 'query', 'sql': count_sql, 'params': []})
            geom_res = pg_tool.call({'operation': 'query', 'sql': geom_sql, 'params': []})
            geom_valid_res = pg_tool.call({'operation': 'query', 'sql': geom_valid_sql, 'params': []})
            return {
                "matched_total": (count_res.get("data") or [{}])[0].get("cnt") if count_res.get("success") else None,
                "geom_total": (geom_res.get("data") or [{}])[0].get("cnt") if geom_res.get("success") else None,
                "geom_valid_total": (geom_valid_res.get("data") or [{}])[0].get("cnt") if geom_valid_res.get("success") else None,
                "lonlat_total": (lonlat_res.get("data") or [{}])[0].get("cnt") if lonlat_res.get("success") else None,
                "matched_total_error": None if count_res.get("success") else count_res.get("error"),
                "geom_total_error": None if geom_res.get("success") else geom_res.get("error"),
                "geom_valid_total_error": None if geom_valid_res.get("success") else geom_valid_res.get("error"),
                "lonlat_total_error": None if lonlat_res.get("success") else lonlat_res.get("error"),
                "where_clause": where_clause,
            }

        sql = f"""
        SELECT json_build_object(
            'type', 'FeatureCollection',
            'features', COALESCE(
                json_agg(
                    json_build_object(
                        'type', 'Feature',
                        'geometry', ST_AsGeoJSON(
                            CASE 
                                WHEN {geom_col} IS NOT NULL THEN 
                                    CASE 
                                        WHEN ST_SRID({geom_col}) = 0 THEN ST_SetSRID(ST_MakeValid({geom_col}), 4326)
                                        ELSE ST_MakeValid({geom_col})
                                    END
                                {lonlat_case}                                ELSE NULL
                            END, 6
                        )::json,
                        'properties', {props_sql}
                    )
                ), 
                '[]'::json
            )
        ) AS geojson
        FROM {safe_table_name} AS t
        {where_clause};
        """

        res = pg_tool.call({'operation': 'query', 'sql': sql, 'params': []})
        if not res.get('success'):
            error_msg = res.get('error', '数据库查询失败')
            logger.warning(f"Vector API query failed for table '{table_name}': {error_msg}")
            # 优雅降级：返回空 FeatureCollection 而非 500，让前端正常处理
            return JSONResponse(
                content={
                    "type": "FeatureCollection",
                    "features": [],
                    "meta": {
                        "status": "error",
                        "message": f"数据表 '{table_name}' 查询失败: {error_msg}",
                        "table_name": table_name,
                        "applied_filter": filter,
                    }
                },
                headers={"Cache-Control": "public, max-age=60"}
            )
        rows = res.get('data') or []

        # 保留失败容错：若两次查询均异常，返回空集合

        if not rows or len(rows) == 0:
            sql2 = f"""
            SELECT 
                ST_AsGeoJSON(
                    CASE 
                        WHEN {geom_col} IS NOT NULL THEN 
                            CASE 
                                WHEN ST_SRID({geom_col}) = 0 THEN ST_SetSRID(ST_MakeValid({geom_col}), 4326)
                                ELSE ST_MakeValid({geom_col})
                            END
                        {lonlat_case}                        ELSE NULL
                    END, 6
                ) AS geom_json,
                (row_to_json(t)::jsonb - '{geom_col}')::json AS props
            FROM {safe_table_name} AS t
            {where_clause};
            """
            res2 = pg_tool.call({'operation': 'query', 'sql': sql2, 'params': []})
            rows2 = res2.get('data') or []
            features2 = []
            for r in rows2:
                gj = r.get("geom_json")
                if not gj:
                    continue
                try:
                    geom = json.loads(gj)
                except:
                    geom = None
                props = r.get("props") or {}
                if geom:
                    features2.append({"type": "Feature", "geometry": geom, "properties": props})
            if features2:
                fc = {"type": "FeatureCollection", "features": features2, "meta": {"feature_count": len(features2), "table_name": table_name, "applied_filter": filter}}
                if debug:
                    fc["_debug"] = {"sql": sql2}
                return JSONResponse(content=fc, headers={"Cache-Control": "public, max-age=60"})
            count_filter = f"({filter})" if filter else "TRUE"
            meta = build_empty_meta(count_filter)
            logger.info(f"Vector query returned no rows: table={table_name}, filter={filter}, meta={meta}")
            content = {
                "type": "FeatureCollection",
                "features": [],
                "meta": {
                    "status": "empty",
                    "message": "查询成功但无可用要素",
                    "applied_filter": filter,
                    "table_name": table_name,
                    **meta
                }
            }
            if debug:
                content["_debug"] = {"sql": sql, **meta}
            return JSONResponse(content=content, headers={"Cache-Control": "public, max-age=60"})

        geojson = rows[0].get('geojson')
        if isinstance(geojson, str):
            try:
                geojson = json.loads(geojson)
            except Exception as e:
                logger.error(f"Vector API returned invalid JSON string: {e}")
                geojson = {"type": "FeatureCollection", "features": [], "meta": {"status": "invalid", "message": "后端返回数据格式异常"}}
        if not isinstance(geojson, dict):
            logger.error(f"Vector API returned non-dict geojson: {type(geojson)}")
            geojson = {"type": "FeatureCollection", "features": [], "meta": {"status": "invalid", "message": "后端返回数据格式异常"}}
        features = geojson.get("features")
        if not isinstance(features, list):
            features = []
            geojson["features"] = features
        feature_count = len(features)
        meta = geojson.get("meta") if isinstance(geojson.get("meta"), dict) else {}
        meta.update({"feature_count": feature_count, "table_name": table_name, "applied_filter": filter})
        geojson["meta"] = meta
        if feature_count == 0:
            count_filter = f"({filter})" if filter else "TRUE"
            empty_meta = build_empty_meta(count_filter)
            meta.update({"status": "empty", "message": "查询成功但无可用要素", **empty_meta})
            logger.info(f"Vector query returned empty features: table={table_name}, filter={filter}, meta={empty_meta}")
            if debug:
                geojson["_debug"] = {"sql": sql, **empty_meta}
        elif debug:
            geojson["_debug"] = {"sql": sql, "where_clause": where_clause}
        return JSONResponse(
            content=geojson,
            headers={"Cache-Control": "public, max-age=60"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Vector API error for table {table_name}")
        raise HTTPException(status_code=500, detail=f"服务器内部错误: {str(e)}")

@app.post("/api/save-screenshot")
async def save_screenshot(payload: ScreenshotRequest):
    if not payload.image_data:
        raise HTTPException(status_code=400, detail="截图数据不能为空")
    match = re.match(r"^data:(image/\w+);base64,", payload.image_data)
    if not match:
        raise HTTPException(status_code=400, detail="无效的图片数据格式")
    mime_type = match.group(1)
    ext = mime_type.split("/")[-1]
    try:
        image_bytes = base64.b64decode(payload.image_data.split(",", 1)[1])
    except Exception:
        raise HTTPException(status_code=400, detail="图片解码失败")
    directory = "/home/server/python/map_assistant_v1/backend/static/screenshots"
    os.makedirs(directory, exist_ok=True)
    base_name = payload.file_name or f"map_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
    safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", base_name)
    if not safe_name.lower().endswith(f".{ext.lower()}"):
        safe_name = f"{safe_name}.{ext}"
    file_path = os.path.join(directory, safe_name)
    with open(file_path, "wb") as f:
        f.write(image_bytes)
    url = f"/static/screenshots/{safe_name}"
    return {"url": url, "filename": safe_name, "file_path": file_path}


# ==============================
# 报告文件强制下载接口
# ==============================
@app.api_route("/api/download/report/{filename}", methods=["GET", "HEAD"])
async def download_report(filename: str):
    """强制触发浏览器下载报告文件（带 Content-Disposition: attachment）"""
    safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", filename)
    file_path = f"/home/server/python/map_assistant_v1/backend/static/reports/{safe_name}"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"报告文件不存在: {safe_name}")
    return FileResponse(
        path=file_path,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'}
    )


# ==============================
# 报告下载接口（无 .docx 后缀，防止迅雷等下载管理器拦截）
# ==============================
@app.get("/api/download/report")
async def download_report_query(filename: str = Query(..., description="报告文件名")):
    """通过 query 参数传递文件名，URL 不含 .docx 后缀，避免迅雷下载管理器拦截"""
    safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", filename)
    file_path = f"/home/server/python/map_assistant_v1/backend/static/reports/{safe_name}"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"报告文件不存在: {safe_name}")
    return FileResponse(
        path=file_path,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'}
    )


# ==============================
# 报告在线预览接口（docx 转 HTML）
# ==============================
@app.get("/api/preview/report", response_class=HTMLResponse)
async def preview_report(filename: str = Query(..., description="报告文件名")):
    """将 docx 报告转为 HTML 在线预览"""
    import mammoth
    safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", filename)
    file_path = f"/home/server/python/map_assistant_v1/backend/static/reports/{safe_name}"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"报告文件不存在: {safe_name}")
    try:
        with open(file_path, "rb") as f:
            result = mammoth.convert_to_html(f)
        html_body = result.value
        warnings = result.messages
        if warnings:
            logger.warning(f"[Preview] mammoth warnings: {warnings[:3]}")
    except Exception as e:
        logger.error(f"[Preview] 转换失败: {e}")
        raise HTTPException(status_code=500, detail=f"报告转换失败: {str(e)}")

    html_page = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>报告预览 - {safe_name}</title>
<style>
  body {{ font-family: "Microsoft YaHei", "PingFang SC", sans-serif; max-width: 860px; margin: 30px auto; padding: 20px 40px; background: #fafafa; color: #333; line-height: 1.8; }}
  h1 {{ font-size: 22px; border-bottom: 2px solid #2563eb; padding-bottom: 8px; color: #1e3a5f; }}
  h2 {{ font-size: 18px; color: #2563eb; margin-top: 24px; }}
  h3 {{ font-size: 15px; color: #475569; }}
  table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
  th, td {{ border: 1px solid #cbd5e1; padding: 8px 12px; text-align: left; font-size: 14px; }}
  th {{ background: #f1f5f9; font-weight: 600; }}
  img {{ max-width: 100%; height: auto; border-radius: 6px; margin: 8px 0; }}
  p {{ margin: 8px 0; }}
  .preview-toolbar {{ position: sticky; top: 0; background: #fff; border-bottom: 1px solid #e2e8f0; padding: 10px 0; margin-bottom: 20px; display: flex; align-items: center; gap: 12px; z-index: 10; }}
  .preview-toolbar h1 {{ font-size: 16px; margin: 0; border: none; padding: 0; flex: 1; }}
  .btn-download {{ background: #2563eb; color: #fff; border: none; border-radius: 6px; padding: 7px 16px; font-size: 13px; cursor: pointer; text-decoration: none; }}
  .btn-download:hover {{ background: #1d4ed8; }}
  .btn-print {{ background: #f1f5f9; color: #334155; border: 1px solid #cbd5e1; border-radius: 6px; padding: 7px 16px; font-size: 13px; cursor: pointer; }}
  .btn-print:hover {{ background: #e2e8f0; }}
  @media print {{ .preview-toolbar {{ display: none; }} body {{ max-width: 100%; padding: 0; }} }}
</style>
</head>
<body>
<div class="preview-toolbar">
  <h1>📄 {safe_name}</h1>
  <button class="btn-print" onclick="window.print()">🖨️ 打印</button>
  <a class="btn-download" href="/api/download/report/{safe_name}" download="{safe_name}">⬇ 下载</a>
</div>
<div class="report-content">
{html_body}
</div>
</body>
</html>"""
    return HTMLResponse(content=html_page)


# ==============================
# 图片上传接口（用于聊天中的多模态图片上传）
# ==============================
UPLOAD_IMAGES_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads")
os.makedirs(UPLOAD_IMAGES_DIR, exist_ok=True)

@app.post("/api/upload_image")
async def upload_image(file: UploadFile = File(...)):
    """上传图片或 Excel 文件，返回访问 URL"""
    import time
    # 校验文件类型（图片 + Excel）
    allowed_types = {
        "image/jpeg", "image/png", "image/gif", "image/webp", "image/bmp",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        "application/vnd.ms-excel",  # .xls
        "application/octet-stream",  # 某些浏览器对 Excel 的兼容类型
    }
    # 也通过后缀名判断，兼容浏览器发送的各种 MIME
    ext = os.path.splitext(file.filename or "image.png")[1].lower()
    is_excel = ext in {".xlsx", ".xls"}
    is_image = ext in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
    if not is_image and not is_excel:
        raise HTTPException(status_code=400, detail=f"不支持的文件格式: {ext}")
    
    # 生成唯一文件名
    ext = os.path.splitext(file.filename or "image.png")[1] or ".png"
    safe_ext = ext.lower() if ext.lower() in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".xlsx", ".xls"} else ".png"
    unique_name = f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}{safe_ext}"
    save_path = os.path.join(UPLOAD_IMAGES_DIR, unique_name)
    
    # 保存文件
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)
    
    # 返回可访问的 URL
    url = f"/static/uploads/{unique_name}"
    logger.info(f"Image uploaded: {url} ({len(content)} bytes)")
    return JSONResponse({"url": url, "filename": unique_name, "size": len(content)})


# ==============================
# SHP 文件上传接口（ZIP 包，解压后供 QGIS MCP 空间分析使用）
# ==============================
SHP_UPLOAD_DIR = "/home/server/python/GIS/uploads"
os.makedirs(SHP_UPLOAD_DIR, exist_ok=True)

REQUIRED_SHP_EXTENSIONS = {".shp", ".dbf", ".shx"}

@app.post("/api/upload/shp")
async def upload_shp(file: UploadFile = File(...)):
    """上传 SHP 文件（ZIP 压缩包），解压到 GIS/uploads/ 目录供 QGIS MCP 使用。
    
    要求 ZIP 至少包含 .shp、.dbf、.shx 三个文件。
    返回图层名称和容器内路径，可直接用于 qgis_mcp_tool 空间分析。
    """
    import time, zipfile, tempfile, shutil

    # 1. 校验文件类型
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext != ".zip":
        raise HTTPException(status_code=400, detail="SHP 文件必须打包为 ZIP 上传（.zip）")

    # 2. 读取并保存到临时文件
    content = await file.read()
    if len(content) > 200 * 1024 * 1024:  # 200MB 限制
        raise HTTPException(status_code=400, detail="文件过大，限制 200MB")

    tmp_zip = os.path.join(tempfile.gettempdir(), f"shp_upload_{uuid.uuid4().hex}.zip")
    try:
        with open(tmp_zip, "wb") as f:
            f.write(content)

        # 3. 验证 ZIP 内容
        with zipfile.ZipFile(tmp_zip, "r") as zf:
            names = zf.namelist()
            exts_in_zip = set()
            shp_stems = set()

            for name in names:
                # 跳过目录和 __MACOSX 隐藏文件
                base = os.path.basename(name)
                if not base or base.startswith("._") or name.endswith("/"):
                    continue
                file_ext = os.path.splitext(base)[1].lower()
                file_stem = os.path.splitext(base)[0].lower()
                if file_ext in REQUIRED_SHP_EXTENSIONS:
                    exts_in_zip.add(file_ext)
                    shp_stems.add(file_stem)

            if not (exts_in_zip >= REQUIRED_SHP_EXTENSIONS):
                missing = REQUIRED_SHP_EXTENSIONS - exts_in_zip
                raise HTTPException(
                    status_code=400,
                    detail=f"ZIP 缺少必要的 SHP 文件: {', '.join(sorted(missing))}（需要 .shp + .dbf + .shx）"
                )

            if len(shp_stems) > 1:
                raise HTTPException(
                    status_code=400,
                    detail=f"ZIP 包含多个 SHP 数据集 ({', '.join(sorted(shp_stems))})，请每次上传一个"
                )

            # 4. 解压到目标目录
            ts = int(time.time())
            uid = uuid.uuid4().hex[:8]
            dest_dir = os.path.join(SHP_UPLOAD_DIR, f"{ts}_{uid}")
            os.makedirs(dest_dir, exist_ok=True)

            shp_name = None
            files_extracted = []
            for name in zf.namelist():
                base = os.path.basename(name)
                if not base or base.startswith("._") or name.endswith("/"):
                    continue
                dest_path = os.path.join(dest_dir, base)
                with zf.open(name) as src:
                    with open(dest_path, "wb") as dst:
                        dst.write(src.read())
                files_extracted.append(base)
                if os.path.splitext(base)[1].lower() == ".shp":
                    shp_name = base

            if not shp_name:
                raise HTTPException(status_code=500, detail="解压后未找到 .shp 文件")

        # 5. 用 ogrinfo 获取图层元数据
        shp_path = os.path.join(dest_dir, shp_name)
        feature_count = 0
        geom_type = "Unknown"
        fields = []
        try:
            import subprocess
            result = subprocess.run(
                ["ogrinfo", "-al", "-so", shp_path],
                capture_output=True, text=True, timeout=10
            )
            for line in result.stdout.split("\n"):
                line = line.strip()
                if line.startswith("Feature Count:"):
                    feature_count = int(line.split(":")[1].strip())
                if line.startswith("Geometry:"):
                    geom_type = line.split(":")[1].strip()
        except Exception as e:
            logger.warning(f"ogrinfo 失败，跳过元数据提取: {e}")

        # 容器内路径
        container_path = f"/uploads/{ts}_{uid}/{shp_name}"
        layer_name = os.path.splitext(shp_name)[0]

        logger.info(f"SHP uploaded: {shp_name} → {dest_dir} (容器: {container_path}, {feature_count} 要素)")

        return JSONResponse({
            "success": True,
            "layer_name": layer_name,
            "container_path": container_path,
            "feature_count": feature_count,
            "geometry_type": geom_type,
            "files": files_extracted,
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"SHP 上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")
    finally:
        # 清理临时 zip
        if os.path.exists(tmp_zip):
            os.unlink(tmp_zip)


# ==============================
# GeoJSON 文件读取接口（供前端矢量图层加载使用）
# ==============================
GEOJSON_DIR = os.path.join(os.path.dirname(__file__), "static", "geojson")

@app.get("/api/geojson/{filename}")
async def get_geojson(filename: str):
    """读取 GeoJSON 文件，用于矢量图层加载到地图"""
    import re
    # 安全校验：仅允许 .geojson 后缀，防止路径穿越
    if not re.match(r'^[a-zA-Z0-9_\-\.]+\.geojson$', filename):
        raise HTTPException(status_code=400, detail="无效的文件名")
    file_path = os.path.join(GEOJSON_DIR, filename)
    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="GeoJSON 文件不存在")
    return FileResponse(file_path, media_type="application/geo+json")


# ==============================
# 原有聊天接口（保持不变）
# ==============================
@app.post("/chat")
async def chat(request: ChatRequest, x_session_id: Optional[str] = Header(default=None, alias="X-Session-ID")):
    try:
        if not request.messages:
            raise HTTPException(status_code=400, detail="消息不能为空")

        # 优先使用请求头 X-Session-ID，其次 body.session_id，最后回退 "default"
        thread_id = x_session_id or request.session_id or "default"
        
        valid_roles = {'user', 'assistant', 'system', 'function'}
        for i, msg in enumerate(request.messages):
            if msg.role not in valid_roles:
                raise HTTPException(status_code=400, detail=f"无效角色: {msg.role}")
        
        messages = [msg.model_dump() for msg in request.messages]
        # 记录每条消息的角色，用于调试 400 错误
        roles = [m.get('role') for m in messages]
        logger.info(f"Processing messages with roles: {roles}")
        
        # 修复：确保消息列表以 user 开头（跳过 system）
        # 如果第一条不是 system 且第一条不是 user，或者第一条是 system 且第二条不是 user
        if messages:
            first_non_system_idx = 0
            if messages[0].get('role') == 'system':
                first_non_system_idx = 1
            
            if len(messages) > first_non_system_idx:
                if messages[first_non_system_idx].get('role') != 'user':
                    logger.warning(f"First non-system message is {messages[first_non_system_idx].get('role')}, not 'user'. Attempting to fix...")
                    # 找到第一个 user 消息并删除它之前的所有非 system 消息
                    first_user_idx = -1
                    for i in range(first_non_system_idx, len(messages)):
                        if messages[i].get('role') == 'user':
                            first_user_idx = i
                            break
                    
                    if first_user_idx != -1:
                        messages = messages[:first_non_system_idx] + messages[first_user_idx:]
                        logger.info(f"Fixed message sequence. New roles: {[m.get('role') for m in messages]}")
                    else:
                        logger.error("No user message found in history!")

        # 强力指令：彻底纠正实测高程与超深逻辑，严禁误导
        if messages and messages[-1]['role'] == 'user':
            content = messages[-1]['content']
            if any(k in content for k in ["超深", "高程", "深度", "采深", "开采深度"]):
                # SSoT：高程业务规则唯一权威在 prompts.py::build_elevation_injection
                messages[-1]['content'] += build_elevation_injection()
                logger.info(f"Injected elevation business rules.")
            # 地图数据加载意图守卫：根据 active_view 注入 2D/3D 视图提示词（SSoT：prompts.py）
            map_intent_keywords = ["加载", "上图", "矢量", "图层", "地图", "位置", "显示到地图", "加载到地图", "跳转", "定位", "标记", "标点", "经纬度", "切换", "卫星", "底图", "清除"]
            if any(k in content for k in map_intent_keywords):
                messages.insert(0, {
                    'role': 'system',
                    'content': build_view_system_message(request.active_view),
                })
                logger.info(f"{'3D' if request.active_view == 'cesium' else '2D'} view detected: view prompt injected.")

        use_intent_agent = os.environ.get("USE_INTENT_AGENT", "true").lower() == "true"

        if use_intent_agent and task_executor is not None:
            result = await task_executor.execute(
                user_message=messages[-1]['content'],
                chat_history=messages[:-1],
                thread_id=thread_id,
            )

            intent_info = None
            if result.get("intent_result"):
                ir = result["intent_result"]
                intent_info = {
                    "primary_intent": ir.primary_intent.value if hasattr(ir.primary_intent, 'value') else str(ir.primary_intent),
                    "confidence": ir.confidence,
                    "task_context": ir.task_context,
                    "entities": ir.entities,
                    "execution_plan": [
                        {
                            "step_id": s.step_id,
                            "action": s.action,
                            "tool": s.tool,
                            "reasoning": s.reasoning
                        }
                        for s in ir.execution_plan
                    ] if ir.execution_plan else []
                }

            # 持久化 intent agent 结果
            _persist_chat(thread_id, messages[-1]['content'], result.get("response", ""))
            _schedule_fact_extraction(thread_id, messages[-1]['content'], result.get("response", ""))

            optimized_map_commands = _optimize_map_commands(messages[-1]['content'], result.get("map_commands", []))
            optimized_charts = _optimize_charts(result.get("charts", []))

            return ChatResponse(
                response=result.get("response", "命令已执行。"),
                messages=result.get("messages", []),
                map_commands=optimized_map_commands,
                cesium_commands=result.get("cesium_commands", []),
                charts=optimized_charts,
                report_url=result.get("report_url"),
                intent_info=intent_info
            )

        response_messages = []
        iteration = 0
        runner = bot
        if messages and messages[-1]['role'] == 'user':
            content = messages[-1]['content']
            weather_intent_keywords = ["天气", "气温", "几度", "空气质量", "AQI", "雾霾", "降雨", "下雨", "预报", "带伞", "风力", "风速", "湿度", "紫外线"]
            if any(k in content for k in weather_intent_keywords):
                runner = bot  # 主 bot 已有 weather_tool
                logger.info("Weather intent: using main bot.")
            else:
                map_intent_keywords = ["加载", "上图", "矢量", "图层", "地图", "位置", "显示到地图", "加载到地图", "跳转", "定位", "标记", "标点", "经纬度"]
                if any(k in content for k in map_intent_keywords):
                    # 始终使用主 bot（已注册所有工具），通过系统消息指导工具选择
                    # 不使用 build_assistant_with_tools 的受限 Assistant，避免工具注册问题
                    runner = bot
                    if request.active_view == 'cesium':
                        logger.info("3D view: using main bot with cesium_tool system hint.")
                    else:
                        logger.info("2D view: using main bot with map_tool system hint.")
        for response in runner.run(messages=messages):
            iteration += 1
            response_messages = response
            # 记录每一次迭代，看是否卡在某个工具调用上
            last_msg = response_messages[-1] if response_messages else {}
            role = last_msg.get('role')
            name = last_msg.get('name', 'N/A')
            content = last_msg.get('content', '')
            
            # 记录关键信息
            if role == 'assistant' and 'call' in str(last_msg):
                logger.info(f"Bot iteration {iteration}: Assistant is calling tool: {last_msg}")
            elif role == 'function':
                logger.info(f"Bot iteration {iteration}: Function {name} returned: {str(content)[:200]}...")
            else:
                logger.info(f"Bot iteration {iteration}: role={role}, content={str(content)[:100]}...")
        
        # 获取最后一条助手回复作为 response 字段（兼容旧前端逻辑）
        response_text = ""
        formatted_answer = ""
        for msg in reversed(response_messages):
            if msg.get('role') == 'assistant' and msg.get('content'):
                response_text = clean_response_content(msg['content'])
                break
        # 如果工具返回了标准化答案（如平均超深深度），优先使用
        if not response_text:
            for msg in reversed(response_messages):
                if msg.get('role') == 'function' and msg.get('name') == 'data_visualizer_tool':
                    try:
                        func_res = json.loads(msg.get('content', '{}'))
                        fa = func_res.get('formatted_answer')
                        if fa:
                            formatted_answer = fa
                            break
                    except:
                        pass
        if formatted_answer:
            response_text = formatted_answer
        
        if not response_text:
            response_text = "命令已执行。"

        # 提取地图命令
        map_commands = []
        cesium_commands = []
        charts = []
        for msg in response_messages:
            if msg.get('role') == 'function' and msg.get('name') in ['map_tool', 'location_search']:
                try:
                    func_res = json.loads(msg.get('content', '{}'))
                    if func_res.get('map_command'):
                        map_commands.append(func_res['map_command'])
                except:
                    pass
            # 提取 Cesium 3D 命令
            if msg.get('role') == 'function' and msg.get('name') == 'cesium_tool':
                try:
                    func_res = json.loads(msg.get('content', '{}'))
                    if func_res.get('cesium_command'):
                        cesium_commands.append(func_res['cesium_command'])
                except:
                    pass
            # 提取数据可视化图表
            if msg.get('role') == 'function' and msg.get('name') == 'data_visualizer_tool':
                try:
                    vis = json.loads(msg.get('content', '{}'))
                    if isinstance(vis, dict) and vis.get('success'):
                        charts.append({
                            'chart_type': vis.get('chart_type'),
                            'config': vis.get('config'),
                            'summary': vis.get('content')
                        })
                except Exception as e:
                    logger.warning(f"Failed to parse visualization content: {e}")
        
        map_commands = _optimize_map_commands(messages[-1]['content'], map_commands)
        charts = _optimize_charts(charts)

        logger.info(f"Chat response: {response_text[:100]}... Map commands: {len(map_commands)}, Cesium commands: {len(cesium_commands)}, Charts: {len(charts)}")
        if map_commands:
            logger.info(f"First map command: {map_commands[0]}")
        if cesium_commands:
            logger.info(f"First cesium command: {cesium_commands[0]}")

        # 持久化消息到 SQLite（仅当 thread_id 是有效 UUID 会话时）
        _persist_chat(thread_id, messages[-1]['content'], response_text)
        _schedule_fact_extraction(thread_id, messages[-1]['content'], response_text)

        return ChatResponse(
            response=response_text,
            messages=response_messages, # 返回完整的对话历史，包括工具执行结果
            map_commands=map_commands,
            cesium_commands=cesium_commands,
            charts=charts
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Chat error: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="内部服务器错误")

# run SSE 订阅空闲超时：超过该时长无事件则检查 run 状态（防哨兵丢失导致挂起）
RUN_SSE_IDLE_TIMEOUT = 120


@app.post("/chat/stream")
async def chat_stream(request: ChatRequest, x_session_id: Optional[str] = Header(default=None, alias="X-Session-ID"), x_pending_run_id: Optional[str] = Header(default=None, alias="X-Pending-Run-ID")):
    def sse_payload(payload: Dict[str, Any]) -> str:
        return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

    async def event_generator():
        try:
            # 先发送 SSE 预热包，尽量让浏览器和代理尽快进入流式模式
            yield ": stream-open\n\n"
            yield f": {' ' * 2048}\n\n"
            yield "retry: 1500\n\n"

            if not request.messages:
                raise HTTPException(status_code=400, detail="消息不能为空")

            thread_id = x_session_id or request.session_id or "default"
            valid_roles = {'user', 'assistant', 'system', 'function'}
            for msg in request.messages:
                if msg.role not in valid_roles:
                    raise HTTPException(status_code=400, detail=f"无效角色: {msg.role}")

            messages = [msg.model_dump() for msg in request.messages]
            roles = [m.get('role') for m in messages]
            logger.info(f"[stream] Processing messages with roles: {roles}")

            # 阶段4 契约改造：前端只传 session_id + 本轮消息时，历史由后端从 DB 读取并裁剪
            # （≤COMPRESS_THRESHOLD_TURNS 轮），避免每轮全量回传造成上下文污染。
            if (len(messages) == 1 and messages[0].get('role') == 'user'
                    and thread_id and thread_id != "default"):
                db_history = load_history_from_db(DB_PATH, thread_id)
                if db_history:
                    # DB 历史与本轮消息不重叠（_persist_chat 在 final 时才写入）
                    messages = db_history + messages
                    logger.info(f"[stream] history loaded from db: {len(db_history)} msgs (session={thread_id})")

            # 阶段4：超长会话滚动摘要压缩（fire-and-forget，失败静默降级为截断）
            if task_executor is not None and getattr(task_executor, "context_manager", None) is not None:
                try:
                    asyncio.create_task(task_executor.context_manager.compress_history(
                        thread_id, messages, getattr(task_executor, "_llm", None),
                    ))
                except Exception as e:
                    logger.warning(f"[stream] compress_history task failed: {e}")

            # 图片 OCR 预处理：用 qwen-vl-ocr 提取文字后注入到用户消息
            # 主模型 qwen-flash 不支持图片，所以文字提取后再移除 images 字段
            messages = await _preprocess_message_images(messages)

            if messages:
                first_non_system_idx = 0
                if messages[0].get('role') == 'system':
                    first_non_system_idx = 1

                if len(messages) > first_non_system_idx and messages[first_non_system_idx].get('role') != 'user':
                    first_user_idx = -1
                    for i in range(first_non_system_idx, len(messages)):
                        if messages[i].get('role') == 'user':
                            first_user_idx = i
                            break
                    if first_user_idx != -1:
                        messages = messages[:first_non_system_idx] + messages[first_user_idx:]

            raw_user_text = None
            if messages and messages[-1]['role'] == 'user':
                content = _extract_text_content(messages[-1]['content'])
                # 保存注入前的原始用户文本：高程业务规则注入会向消息追加含快速路由
                # 关键词的文本（如"红线标准"），快速路由/意图分析必须基于纯用户输入，
                # 否则"XX可采区的高程是多少"会被注入文本误导为 spatial_reference。
                raw_user_text = content
                if any(k in content for k in ["超深", "高程", "深度", "采深", "开采深度"]):
                    # 多模态消息需要特殊处理：将业务规则追加到文本部分
                    # SSoT：高程业务规则唯一权威在 prompts.py::build_elevation_injection
                    extra_rules = build_elevation_injection()
                    if isinstance(messages[-1]['content'], list):
                        # 多模态格式：追加到第一个 text 部分
                        for part in messages[-1]['content']:
                            if isinstance(part, dict) and part.get("type") == "text":
                                part["text"] = part.get("text", "") + extra_rules
                                break
                    else:
                        messages[-1]['content'] += extra_rules
                map_intent_keywords = ["加载", "上图", "矢量", "图层", "地图", "位置", "显示到地图", "加载到地图", "跳转", "定位", "标记", "标点", "经纬度", "切换", "卫星", "底图", "清除"]
                if any(k in content for k in map_intent_keywords):
                    # 视图提示词 SSoT：prompts.py::build_view_system_message（阶段4/6）
                    messages.insert(0, {
                        'role': 'system',
                        'content': build_view_system_message(request.active_view),
                    })

            yield sse_payload({
                "type": "status",
                "stage": "queued",
                "message": "请求已提交，后端开始处理。",
            })

            use_intent_agent = os.environ.get("USE_INTENT_AGENT", "true").lower() == "true"
            run_engine_on = os.environ.get("RUN_ENGINE", "on").lower() in ("1", "on", "true")

            # ── 阶段1新路径：RunEngine（灰度开关 RUN_ENGINE，默认 on）──
            # run 在独立 asyncio task 中执行，SSE 只是事件订阅者；断线不影响执行
            if run_engine_on and task_executor is not None and task_executor.run_engine is not None:
                user_text = raw_user_text or _extract_text_content(messages[-1].get('content', ''))
                engine = task_executor.run_engine
                store = engine.store
                bus = engine.bus

                # pending resume 判定（规则化：确认 / 补参，不调 LLM）
                run_id = None
                task = None
                pending = {}
                pending_run = None
                resume_source_text = None  # resume 场景的原始请求文本（用于地图命令优化判定）
                if x_pending_run_id:
                    pending_run = store.get_run(x_pending_run_id)
                    if pending_run and pending_run.get("status") in ("awaiting_confirmation", "awaiting_input"):
                        try:
                            pending = json.loads(pending_run.get("pending_json") or "{}")
                        except Exception:
                            pending = {}
                    else:
                        pending_run = None
                if pending_run is None:
                    pending_run = store.get_pending_by_session(thread_id)
                    pending = (pending_run or {}).get("pending") or {}

                if pending_run:
                    rid = pending_run["run_id"]
                    ptype = (pending or {}).get("pending_type", "")
                    if ptype == "confirm" and is_confirm_message(user_text):
                        task = await engine.resume(rid, confirm=True)
                        run_id = rid
                        logger.info(f"[stream] resume confirm run={rid}")
                    elif ptype == "input":
                        supplied = parse_supplied_from_message(pending, user_text)
                        if supplied:
                            task = await engine.resume(rid, user_supplied=supplied)
                            run_id = rid
                            logger.info(f"[stream] resume input run={rid} supplied={list(supplied.keys())}")
                    if task is not None:
                        # 补参/确认消息本身不含地图关键词（如"ceshen"），但原始请求
                        # 已判定为地图意图；用 checkpoint 中的原始 user_message 做
                        # 地图命令优化判定，避免 resume 场景误抑制 map_commands。
                        ckpt = store.load_checkpoint(rid)
                        resume_source_text = (ckpt or {}).get("user_message") or None
                    else:
                        logger.info(f"[stream] pending run={rid} 未匹配 resume 规则，另起新 run")

                if task is None:
                    run_id = str(uuid.uuid4())
                    task = await engine.start(
                        run_id=run_id, session_id=thread_id, user_message=user_text,
                        chat_history=messages[:-1],
                        view_hint=request.active_view,
                    )

                yield sse_payload({
                    "type": "status", "stage": "run_started",
                    "message": "请求已提交，run 已启动。",
                    "run_id": run_id,
                })

                # 订阅事件流并转 SSE。known_seq 为订阅前的历史边界：
                # 回放中的旧 final（如上一生命周期 pending 终结事件，resume 场景）
                # 不应终止本次 SSE，跳过它们等待真正的最终结果。
                known_seq = await bus.current_seq(run_id)
                queue = await bus.subscribe(run_id)
                try:
                    while True:
                        try:
                            event = await asyncio.wait_for(queue.get(), timeout=RUN_SSE_IDLE_TIMEOUT)
                        except asyncio.TimeoutError:
                            run = store.get_run(run_id)
                            if run and run["status"] in ("completed", "failed", "cancelled"):
                                break
                            continue
                        if event is None:
                            break
                        if event.get("type") == "final":
                            result = event.get("result", {})
                            if result.get("pending_type") and event.get("seq", 0) <= known_seq:
                                # 回放的旧 pending 终结事件：跳过，等新生命周期的事件
                                continue
                            result["run_id"] = run_id
                            result["map_commands"] = _optimize_map_commands(
                                resume_source_text or user_text, result.get("map_commands", [])
                            )
                            result["charts"] = _optimize_charts(result.get("charts", []))
                            _persist_chat(thread_id, user_text, result.get("response", ""))
                            _schedule_fact_extraction(thread_id, user_text, result.get("response", ""))
                            yield sse_payload({"type": "final", "result": result})
                            break
                        yield sse_payload(event)
                finally:
                    bus.unsubscribe(run_id, queue)

                yield "data: [DONE]\n\n"
                return

            if use_intent_agent and task_executor is not None:
                # 旧路径：execute_stream 手动编排（RUN_ENGINE=off 时保留）
                # 提取纯文本用于 task_executor（它目前只支持字符串 user_message）
                user_text = _extract_text_content(messages[-1].get('content', ''))
                async for event in task_executor.execute_stream(
                    user_message=user_text,
                    chat_history=messages[:-1],
                    thread_id=thread_id,
                ):
                    if event.get("type") == "final":
                        result = event.get("result", {})
                        result["map_commands"] = _optimize_map_commands(user_text, result.get("map_commands", []))
                        result["charts"] = _optimize_charts(result.get("charts", []))
                        _persist_chat(thread_id, user_text, result.get("response", ""))
                        _schedule_fact_extraction(thread_id, user_text, result.get("response", ""))
                        yield sse_payload({"type": "final", "result": result})
                    else:
                        yield sse_payload(event)

                yield "data: [DONE]\n\n"
                return

            response_messages = []
            iteration = 0
            runner = bot
            if messages and messages[-1]['role'] == 'user':
                content = messages[-1]['content']
                weather_intent_keywords = ["天气", "气温", "几度", "空气质量", "AQI", "雾霾", "降雨", "下雨", "预报", "带伞", "风力", "风速", "湿度", "紫外线"]
                if any(k in content for k in weather_intent_keywords):
                    runner = bot
                else:
                    map_intent_keywords = ["加载", "上图", "矢量", "图层", "地图", "位置", "显示到地图", "加载到地图", "跳转", "定位", "标记", "标点", "经纬度"]
                    if any(k in content for k in map_intent_keywords):
                        runner = bot

            yield sse_payload({
                "type": "status",
                "stage": "model",
                "message": "模型已开始执行，正在逐步调用工具。",
            })

            for response in runner.run(messages=messages):
                iteration += 1
                response_messages = response
                last_msg = response_messages[-1] if response_messages else {}
                role = last_msg.get('role')
                name = last_msg.get('name', 'N/A')
                content = last_msg.get('content', '')

                if role == 'assistant' and 'call' in str(last_msg):
                    yield sse_payload({
                        "type": "tool_start",
                        "stage": "tool_start",
                        "tool_name": name,
                        "message": f"正在调用工具处理请求（第 {iteration} 轮）。",
                    })
                elif role == 'function':
                    try:
                        parsed = json.loads(content) if isinstance(content, str) else content
                    except Exception:
                        parsed = {"content": str(content)}
                    summary = parsed.get('content') or parsed.get('message') or parsed.get('error') or f"{name} 已返回结果"
                    yield sse_payload({
                        "type": "tool_result",
                        "stage": "tool_result",
                        "tool_name": name,
                        "message": f"{name}：{str(summary)[:100]}",
                    })
                elif role == 'assistant' and content:
                    preview = clean_response_content(str(content))[:80]
                    if preview:
                        yield sse_payload({
                            "type": "status",
                            "stage": "reasoning",
                            "message": f"正在整理回复：{preview}",
                        })

            response_text = ""
            formatted_answer = ""
            for msg in reversed(response_messages):
                if msg.get('role') == 'assistant' and msg.get('content'):
                    response_text = clean_response_content(msg['content'])
                    break
            if not response_text:
                for msg in reversed(response_messages):
                    if msg.get('role') == 'function' and msg.get('name') == 'data_visualizer_tool':
                        try:
                            func_res = json.loads(msg.get('content', '{}'))
                            fa = func_res.get('formatted_answer')
                            if fa:
                                formatted_answer = fa
                                break
                        except Exception:
                            pass
            if formatted_answer:
                response_text = formatted_answer
            if not response_text:
                response_text = "命令已执行。"

            map_commands = []
            cesium_commands = []
            charts = []
            for msg in response_messages:
                if msg.get('role') == 'function' and msg.get('name') in ['map_tool', 'location_search']:
                    try:
                        func_res = json.loads(msg.get('content', '{}'))
                        if func_res.get('map_command'):
                            map_commands.append(func_res['map_command'])
                    except Exception:
                        pass
                if msg.get('role') == 'function' and msg.get('name') == 'cesium_tool':
                    try:
                        func_res = json.loads(msg.get('content', '{}'))
                        if func_res.get('cesium_command'):
                            cesium_commands.append(func_res['cesium_command'])
                    except Exception:
                        pass
                if msg.get('role') == 'function' and msg.get('name') == 'data_visualizer_tool':
                    try:
                        vis = json.loads(msg.get('content', '{}'))
                        if isinstance(vis, dict) and vis.get('success'):
                            charts.append({
                                'chart_type': vis.get('chart_type'),
                                'config': vis.get('config'),
                                'summary': vis.get('content')
                            })
                    except Exception as e:
                        logger.warning(f"[stream] Failed to parse visualization content: {e}")

            map_commands = _optimize_map_commands(messages[-1]['content'], map_commands)
            charts = _optimize_charts(charts)
            _persist_chat(thread_id, messages[-1]['content'], response_text)
            _schedule_fact_extraction(thread_id, messages[-1]['content'], response_text)

            yield sse_payload({
                "type": "final",
                "result": {
                    "response": response_text,
                    "messages": response_messages,
                    "map_commands": map_commands,
                    "cesium_commands": cesium_commands,
                    "charts": charts,
                    "intent_info": None,
                }
            })
            yield "data: [DONE]\n\n"
        except HTTPException as e:
            message = e.detail if isinstance(e.detail, str) else "请求处理失败"
            # 失败轮次也落库，避免 DB 历史断档
            if messages:
                _persist_chat(thread_id, messages[-1]['content'], message)
            yield sse_payload({"type": "error", "stage": "error", "message": message})
            yield sse_payload({
                "type": "final",
                "result": {
                    "response": message,
                    "messages": [],
                    "map_commands": [],
                    "cesium_commands": [],
                    "charts": [],
                    "intent_info": None,
                }
            })
            yield "data: [DONE]\n\n"
        except Exception as e:
            logger.error(f"Chat stream error: {e}\n{traceback.format_exc()}")
            message = f"内部服务器错误：{str(e)}"
            # 失败轮次也落库，避免 DB 历史断档
            if messages:
                _persist_chat(thread_id, messages[-1]['content'], message)
            yield sse_payload({"type": "error", "stage": "error", "message": message})
            yield sse_payload({
                "type": "final",
                "result": {
                    "response": message,
                    "messages": [],
                    "map_commands": [],
                    "cesium_commands": [],
                    "charts": [],
                    "intent_info": None,
                }
            })
            yield "data: [DONE]\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
            "Content-Type": "text/event-stream; charset=utf-8",
            "Content-Encoding": "identity",
        },
    )


# ---------------------------------------------------------------------------
# 阶段1：Run 生命周期端点（查询 / 断线补拉 / 取消）
# ---------------------------------------------------------------------------

@app.get("/api/run/{run_id}")
async def get_run_status(run_id: str):
    """查询 run 状态与 pending 载荷（断线重连 / 前端轮询）。"""
    store = get_run_store()
    run = store.get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail=f"run {run_id} 不存在")
    out = {k: run[k] for k in ("run_id", "session_id", "status", "user_message", "created_at", "updated_at")}
    pending_json = run.get("pending_json")
    out["pending"] = json.loads(pending_json) if pending_json else None
    return out


@app.get("/api/run/{run_id}/events")
async def get_run_events(run_id: str, since: int = Query(default=0, ge=0)):
    """断线补拉：返回 seq > since 的事件（升序）。内存缓冲优先，回退 DB。"""
    store = get_run_store()
    if not store.get_run(run_id):
        raise HTTPException(status_code=404, detail=f"run {run_id} 不存在")
    bus = None
    if task_executor is not None and task_executor.run_engine is not None:
        bus = task_executor.run_engine.bus
    if bus is not None:
        events = await bus.get_history(run_id, since)
    else:
        events = store.get_events(run_id, since)
    latest_seq = max([e.get("seq", 0) for e in events] or [since])
    return {"run_id": run_id, "events": events, "latest_seq": latest_seq}


@app.post("/api/run/{run_id}/cancel")
async def cancel_run(run_id: str):
    """取消 run：置取消标记，引擎在下一个步骤中断点响应并停止（结果丢弃不写 workspace）。"""
    store = get_run_store()
    run = store.get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail=f"run {run_id} 不存在")
    if run["status"] in ("completed", "failed", "cancelled"):
        return {"run_id": run_id, "status": run["status"], "cancelled": False,
                "message": "run 已处于终态，无需取消"}
    store.set_cancelled(run_id)
    logger.info(f"[run] cancel requested: {run_id}")
    return {"run_id": run_id, "status": "cancelled", "cancelled": True,
            "message": "取消标记已设置，run 将在当前步骤完成后停止"}

from tools.overlay_tile_service import get_tile_png as _overlay_get_tile_png, list_layers as _overlay_list_layers, query_feature as _overlay_query_feature, register_layer as _overlay_register_layer, unregister_layer as _overlay_unregister_layer
from tools import geoserver_client as _gs_client
from tools.geoserver_client import GeoServerUnavailable as _GeoServerUnavailable

# GeoLibre 图层工作台默认样式（与 GeoLibre 项目 schema 对齐）
_GEOLIBRE_STYLE = {
    "minZoom": 0, "maxZoom": 24,
    "fillColor": "#1d4ed8", "strokeColor": "#1e3a8a", "strokeWidth": 2, "fillOpacity": 0.25,
    "circleRadius": 6, "textColor": "#111827", "textHaloColor": "#ffffff",
    "textHaloWidth": 2, "textSize": 16,
    "extrusionEnabled": False, "extrusionColor": "#3b82f6", "extrusionOpacity": 0.8,
    "extrusionHeightProperty": "height", "extrusionHeightScale": 1, "extrusionBase": 0,
    "extrusionAdvancedStyleEnabled": False, "extrusionColorExpression": "",
    "extrusionHeightExpression": "", "vectorStyleMode": "single", "vectorStyleProperty": "",
    "vectorStyleClassCount": 5, "vectorStyleColorRamp": "viridis",
    "vectorStyleClassificationScheme": "equal-interval",
    "vectorStyleStops": [{"value": 0, "color": "#dbeafe"}, {"value": 1, "color": "#2563eb"}],
    "vectorStyleExpression": "", "pointRenderer": "single",
    "heatmapRadius": 30, "heatmapIntensity": 1, "clusterRadius": 50, "clusterMaxZoom": 14,
    "rasterBrightnessMin": 0, "rasterBrightnessMax": 1, "rasterSaturation": 0,
    "rasterContrast": 0, "rasterHueRotate": 0,
}


def _compute_3dtiles_ground_offset(tileset_path: str) -> float:
    """估算 3D Tiles 在 GeoLibre（地表为 0m 椭球面）中的贴地 altitudeOffset（米，负值=下移）。

    maplibre-gl-3d-tiles 以根节点包围体中心为锚点放置模型，因此
    偏移量 = -(包围体中心海拔 - 包围体垂直半高)，将包围体底面压到 0m 地表。
    """
    try:
        from pyproj import Transformer

        with open(tileset_path, "r", encoding="utf-8") as fh:
            t = json.load(fh)
        root = t.get("root") or {}
        tr = root.get("transform")
        if not tr or len(tr) < 12:
            return 0.0
        tx, ty, tz = tr[12], tr[13], tr[14]
        n = math.sqrt(tx * tx + ty * ty + tz * tz) or 1.0
        up = (tx / n, ty / n, tz / n)
        bv = root.get("boundingVolume") or {}
        if "box" in bv and len(bv["box"]) == 12:
            b = bv["box"]
            # 包围盒世界中心 = R @ local_center + t（transform 为 column-major）
            cx = tr[0] * b[0] + tr[4] * b[1] + tr[8] * b[2] + tx
            cy = tr[1] * b[0] + tr[5] * b[1] + tr[9] * b[2] + ty
            cz = tr[2] * b[0] + tr[6] * b[1] + tr[10] * b[2] + tz
            # 垂直半高 = 三个半轴在 up 方向投影绝对值之和
            vhalf = 0.0
            for i, half in enumerate((b[3], b[7], b[11])):
                col = (tr[4 * i], tr[4 * i + 1], tr[4 * i + 2])
                ln = math.sqrt(col[0] ** 2 + col[1] ** 2 + col[2] ** 2) or 1.0
                vhalf += abs(half) * abs(col[0] * up[0] + col[1] * up[1] + col[2] * up[2]) / ln
        elif "sphere" in bv and len(bv["sphere"]) == 4:
            s = bv["sphere"]
            cx = tr[0] * s[0] + tr[4] * s[1] + tr[8] * s[2] + tx
            cy = tr[1] * s[0] + tr[5] * s[1] + tr[9] * s[2] + ty
            cz = tr[2] * s[0] + tr[6] * s[1] + tr[10] * s[2] + tz
            vhalf = abs(s[3])
        else:
            cx, cy, cz = tx, ty, tz
            vhalf = 0.0
        _lon, _lat, alt = Transformer.from_crs(
            "EPSG:4978", "EPSG:4979", always_xy=True
        ).transform(cx, cy, cz)
        bottom = alt - vhalf
        return round(-bottom) if bottom > 0 else 0.0
    except Exception as exc:
        logging.warning("计算 3D Tiles 贴地偏移失败 %s: %s", tileset_path, exc)
        return 0.0


@app.get("/api/geolibre/project")
async def geolibre_project(request: Request):
    """为 GeoLibre 图层工作台动态生成 .geolibre.json 项目。

    预置四个图层：2023年高分影像、河道红线、2026年采区边界、北汝河实景三维。
    host 依据请求动态拼接，保证 GeoLibre(8090) 内各图层 URL 指向当前服务器。
    """
    base = f"{request.url.scheme}://{request.url.netloc}"

    # 读取 2026 年采区边界（geojson 内嵌；数据量小，直接嵌入项目）
    overlay_data_dir = os.environ.get(
        "MAP_OVERLAY_DATA_DIR",
        "/home/server/python/map_assistant_v1/frontend/public/data",
    )
    caiqu2026_path = os.path.join(overlay_data_dir, "caiqu2026.geojson")
    caiqu2026 = {"type": "FeatureCollection", "features": []}
    if os.path.isfile(caiqu2026_path):
        try:
            with open(caiqu2026_path, "r", encoding="utf-8") as fh:
                caiqu2026 = json.load(fh)
        except Exception as exc:
            logging.warning(f"读取 caiqu2026.geojson 失败: {exc}")

    # 北汝河 3D Tiles 贴地偏移：GeoLibre 地表为 0m 椭球面，按包围盒底面下移
    _beiruhe_meta = _load_3dtiles_registry().get("jiaxian-beiruhe") or {}
    _beiruhe_dir = _beiruhe_meta.get("directory") or os.path.join(_3DTILES_DATA_DIR, "jiaxian-beiruhe")
    beiruhe_offset = _compute_3dtiles_ground_offset(os.path.join(_beiruhe_dir, "tileset.json"))

    layers = [
        {
            "id": "geolibre-gf-2023",
            "name": "2023年高分影像",
            "type": "xyz",
            "visible": True,
            "opacity": 1,
            "style": _GEOLIBRE_STYLE,
            "metadata": {"sourceKind": "xyz-url"},
            "source": {
                "type": "raster",
                "tiles": [f"{base}/proxy/gf2023-tiles/{{z}}/{{y}}/{{x}}"],
                "tileSize": 256,
                "url": f"{base}/proxy/gf2023-tiles/{{z}}/{{y}}/{{x}}",
                "attribution": "2023年高分影像",
                "maxzoom": 18,
            },
        },
        {
            "id": "geolibre-hx",
            "name": "河道红线",
            "type": "xyz",
            "visible": True,
            "opacity": 1,
            "style": _GEOLIBRE_STYLE,
            "metadata": {"sourceKind": "xyz-url"},
            "source": {
                "type": "raster",
                "tiles": [f"{base}/api/overlay_tile/hx/{{z}}/{{x}}/{{y}}.png"],
                "tileSize": 256,
                "url": f"{base}/api/overlay_tile/hx/{{z}}/{{x}}/{{y}}.png",
                "attribution": "河道红线",
                "maxzoom": 18,
            },
        },
        {
            "id": "geolibre-caiqu-2026",
            "name": "2026年采区边界",
            "type": "geojson",
            "visible": True,
            "opacity": 1,
            "style": {
                **_GEOLIBRE_STYLE,
                "fillColor": "#1d4ed8",
                "strokeColor": "#1e3a8a",
                "fillOpacity": 0.3,
                "strokeWidth": 2,
            },
            "metadata": {},
            "source": {"type": "geojson"},
            "geojson": caiqu2026,
        },
        {
            "id": "geolibre-beiruhe-3dtiles",
            "name": "北汝河实景三维",
            "type": "3d-tiles",
            "visible": True,
            "opacity": 1,
            "style": _GEOLIBRE_STYLE,
            "metadata": {
                "sourceKind": "3d-tiles-url",
                "externalNativeLayer": True,
                "customLayerType": "3d-tiles",
                "identifiable": False,
            },
            "source": {
                "type": "3d-tiles",
                "url": f"{base}/api/3dtiles/jiaxian-beiruhe/tileset.json",
                "altitudeOffset": beiruhe_offset,
            },
            "sourcePath": f"{base}/api/3dtiles/jiaxian-beiruhe/tileset.json",
        },
    ]

    return {
        "version": "0.1.0",
        "name": "豫水智能一张图 - 图层工作台",
        "mapView": {"center": [114.0, 32.1], "zoom": 11, "bearing": 0, "pitch": 0},
        "basemapStyleUrl": "https://tiles.openfreemap.org/styles/liberty",
        "basemapVisible": True,
        "basemapOpacity": 1,
        "layers": layers,
        "styles": {},
        "preferences": {
            "map": {
                "restrictBounds": False,
                "bounds": [-180, -85, 180, 85],
                "minZoom": 0,
                "maxZoom": 24,
                "maxPitch": 85,
                "renderWorldCopies": True,
            },
            "environmentVariables": [],
        },
        "metadata": {"generated_by": "yushui_map_assistant"},
    }


@app.get("/api/overlay_tile/{layer}/{z}/{x}/{y}.png")
async def overlay_tile(layer: str, z: int, x: int, y: int):
    """矢量图层的栅格瓦片接口（hx / caiqu 等）。"""
    if layer not in _overlay_list_layers():
        raise HTTPException(status_code=404, detail=f"unknown overlay layer: {layer}")
    if z < 0 or z > 22:
        raise HTTPException(status_code=400, detail="invalid zoom")
    n = 1 << z
    if x < 0 or x >= n or y < 0 or y >= n:
        raise HTTPException(status_code=400, detail="invalid tile xy")
    try:
        png_bytes = _overlay_get_tile_png(layer, z, x, y)
    except Exception as e:
        logger.exception("overlay tile render failed: %s/%s/%s/%s", layer, z, x, y)
        raise HTTPException(status_code=500, detail=str(e))
    return Response(
        content=png_bytes,
        media_type="image/png",
        headers={
            "Cache-Control": "public, max-age=86400",
            "Access-Control-Allow-Origin": "*",
        },
    )


_VT_DIR = os.path.join(os.path.dirname(__file__), "vector_tiles")

@app.get("/api/vector_tile/{layer}/{z}/{x}/{y}.pbf")
async def vector_tile(layer: str, z: int, x: int, y: int):
    """预生成的矢量切片接口（tippecanoe .pbf），供 Leaflet 2D 地图使用。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", layer):
        raise HTTPException(status_code=404, detail=f"unknown vector tile layer: {layer}")
    pbf_path = os.path.join(_VT_DIR, layer, str(z), str(x), f"{y}.pbf")
    if not os.path.abspath(pbf_path).startswith(os.path.abspath(_VT_DIR) + os.sep):
        raise HTTPException(status_code=400, detail="invalid tile path")
    if not os.path.isfile(pbf_path):
        return Response(status_code=204)
    return FileResponse(
        pbf_path,
        media_type="application/x-protobuf",
        headers={
            "Cache-Control": "public, max-age=604800",
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.get("/api/overlay_feature/{layer}")
async def overlay_feature(layer: str, lng: float, lat: float, tolerance_m: float = 120.0):
    if layer not in _overlay_list_layers():
        raise HTTPException(status_code=404, detail=f"unknown overlay layer: {layer}")
    if not (-180 <= lng <= 180 and -90 <= lat <= 90):
        raise HTTPException(status_code=400, detail="invalid coordinate")
    try:
        feature = _overlay_query_feature(layer, lng, lat, tolerance_m)
    except Exception as e:
        logger.exception("overlay feature query failed: %s/%s/%s", layer, lng, lat)
        raise HTTPException(status_code=500, detail=str(e))
    return JSONResponse(
        content={"found": feature is not None, "feature": feature},
        headers={"Cache-Control": "no-store", "Access-Control-Allow-Origin": "*"},
    )


class TileRegenerateRequest(BaseModel):
    layer: str


class DroneImageryRegisterRequest(BaseModel):
    layer_key: str
    name: str = ""
    path: str
    area_key: str = ""
    year: Optional[int] = None
    min_zoom: int = 0
    max_zoom: int = 22
    max_native_zoom: Optional[int] = None
    bounds: Optional[List[float]] = None
    opacity: float = 0.9
    scheme: str = "tms"


class DroneImageryBuildRequest(BaseModel):
    source_path: str
    layer_key: str = ""
    name: str = ""
    area_key: str = ""
    year: Optional[int] = None
    min_zoom: int = 0
    max_zoom: int = 22
    opacity: float = 0.9
    tile_format: str = "PNG"
    quality: int = 85
    overwrite: bool = True
    source_srs: str = ""  # 手动指定源坐标系，如 EPSG:4547；为空则自动检测


# ---------- 3D Tiles 管理 ----------
class ThreeDTilesRegisterRequest(BaseModel):
    directory: str
    key: str = ""
    name: str = ""
    label: str = ""
    alt_offset: float = 0.0
    auto_ground_clamp: bool = True
    description: str = ""


class ThreeDTilesDeleteRequest(BaseModel):
    key: str
    delete_files: bool = False


# 在模块加载时自动注册已有数据集（后台线程执行，不阻塞启动）
_auto_register_existing_3dtiles_async()


def _register_custom_raster_layers() -> None:
    for key, meta in _load_tile_registry().items():
        if meta.get("build_type", "both") not in ("raster", "both"):
            continue
        source = meta.get("source")
        style = meta.get("style") or {}
        if source and os.path.isfile(source):
            try:
                _overlay_register_layer(key, source, style)
            except Exception as e:
                logger.warning("register custom raster layer failed %s: %s", key, e)


_register_custom_raster_layers()


# ---------------------------------------------------------------------------
# 后台构建任务（tippecanoe / GeoServer 发布等耗时操作放入线程，避免阻塞服务）
# ---------------------------------------------------------------------------

def _run_geojson_build_job(
    job: Dict[str, Any],
    safe_key: str,
    source_path: str,
    meta: Dict[str, Any],
    build_type: str,
    min_zoom: int,
    max_zoom: int,
    style: Dict[str, Any],
    auto_publish: bool,
) -> None:
    """后台执行 GeoJSON 构建：tippecanoe + 可选 GeoServer 发布。"""
    def _update(**kw):
        job.update(kw)
        job["updated_at"] = dt.now().isoformat(timespec="seconds")

    try:
        stats = {"tile_count": 0, "size_bytes": 0}
        if build_type in ("vector", "both"):
            _update(stage="tippecanoe", percent=5, message="开始生成矢量切片（tippecanoe）...")
            stats = _run_tippecanoe(safe_key, source_path, min_zoom, max_zoom)
            _update(stage="tippecanoe", percent=60, message="矢量切片生成完成")
        geoserver_result = None
        if auto_publish:
            _update(stage="publish", percent=70, message="正在发布到 GeoServer...")
            try:
                import_result = _gs_client.import_geojson_to_postgis(safe_key, source_path)
                geoserver_result = _gs_client.publish_by_tm_key(safe_key, {
                    "type": "vector",
                    "label": meta["label"],
                    "source_path": source_path,
                    "style": style,
                })
                geoserver_result["import"] = import_result
                _update(percent=90, message="GeoServer 发布完成")
            except _GeoServerUnavailable as e:
                logger.warning("GeoServer 自动发布失败: %s", e)
                geoserver_result = {"error": str(e)}
                _update(percent=95, message=f"GeoServer 发布失败（已跳过）: {str(e)[:100]}")
        _invalidate_tile_stats()
        _update(
            stage="done", percent=100, message="构建完成",
            success=True, done=True, stats=stats, geoserver=geoserver_result, layer=safe_key,
        )
    except HTTPException as e:
        _update(stage="error", percent=0, message=str(e.detail), success=False, done=True)
    except Exception as e:
        logger.exception("geojson build job failed: %s", safe_key)
        _update(stage="error", percent=0, message=str(e)[:200], success=False, done=True)


@app.get("/api/tile_manager/layers")
async def tile_manager_layers():
    layers = []
    for key, meta in _merged_tile_meta().items():
        source_path = meta["source"]
        style = meta.get("style") or {}
        min_zoom = int(meta.get("min_zoom", 0))
        max_zoom = int(meta.get("max_zoom", 18))
        build_type = meta.get("build_type", "both")
        vector_dir = os.path.join(_VT_DIR, key)
        stats = _dir_stats(vector_dir)
        vector_ready = build_type in ("vector", "both") and stats["tile_count"] > 0
        raster_ready = build_type in ("raster", "both") and os.path.isfile(source_path)
        color = style.get("stroke") or meta.get("color") or "#2773d7"
        layers.append({
            "key": key,
            "label": meta["label"],
            "type": "vector",
            "status": "ready" if vector_ready else "missing",
            "color": color,
            "tile_count": stats["tile_count"],
            "size_bytes": stats["size_bytes"],
            "min_zoom": min_zoom,
            "max_zoom": max_zoom,
            "api_url": f"/api/vector_tile/{key}/{{z}}/{{x}}/{{y}}.pbf",
            "directory": vector_dir,
            "source_path": source_path,
            "source_name": os.path.basename(source_path),
            "style": style,
            "custom": key not in _TILE_LAYER_META,
        })
        layers.append({
            "key": key,
            "label": meta["label"],
            "type": "raster",
            "status": "ready" if raster_ready else "missing",
            "color": color,
            "tile_count": None,
            "size_bytes": os.path.getsize(source_path) if os.path.isfile(source_path) else 0,
            "min_zoom": 0,
            "max_zoom": 22,
            "api_url": f"/api/overlay_tile/{key}/{{z}}/{{x}}/{{y}}.png",
            "directory": "后端实时渲染 + LRU 缓存",
            "source_path": source_path,
            "source_name": os.path.basename(source_path),
            "style": style,
            "custom": key not in _TILE_LAYER_META,
        })
    for key, meta in _load_drone_registry().items():
        if isinstance(meta, dict):
            layers.append(_drone_layer_to_row(key, meta))
    for key, meta in _load_3dtiles_registry().items():
        if isinstance(meta, dict):
            layers.append(_3dtiles_layer_to_row(key, meta))
    return JSONResponse(content={"success": True, "layers": layers})


@app.post("/api/tile_manager/regenerate")
async def tile_manager_regenerate(req: TileRegenerateRequest):
    key = req.layer
    meta_map = _merged_tile_meta()
    if key not in meta_map:
        raise HTTPException(status_code=404, detail=f"unknown tile layer: {key}")
    source_path = meta_map[key]["source"]
    if not os.path.isfile(source_path):
        raise HTTPException(status_code=404, detail=f"source geojson not found: {source_path}")
    min_zoom = int(meta_map[key].get("min_zoom", 0))
    max_zoom = int(meta_map[key].get("max_zoom", 18))

    def _run(job: Dict[str, Any]) -> None:
        job["stage"] = "tippecanoe"
        job["percent"] = 5
        job["message"] = "开始重新生成矢量切片..."
        job["updated_at"] = dt.now().isoformat(timespec="seconds")
        try:
            stats = _run_tippecanoe(key, source_path, min_zoom, max_zoom)
            _invalidate_tile_stats()
            job.update({
                "stage": "done", "percent": 100, "message": "重新生成完成",
                "success": True, "done": True, "layer": key, **stats,
            })
        except HTTPException as e:
            job.update({
                "stage": "error", "percent": 0, "message": str(e.detail),
                "success": False, "done": True,
            })
        except Exception as e:
            job.update({
                "stage": "error", "percent": 0, "message": str(e)[:200],
                "success": False, "done": True,
            })

    job_id = _submit_tile_build_job({}, _run)
    return JSONResponse(content={"success": True, "layer": key, "async": True, "job_id": job_id})


@app.get("/api/tile_manager/build_status/{job_id}")
async def tile_manager_build_status(job_id: str):
    """查询后台切片构建任务状态。"""
    job = _TILE_BUILD_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="构建任务不存在或已过期")
    return JSONResponse(content={"success": True, "job": job})


@app.delete("/api/tile_manager/{layer_key}")
async def tile_manager_delete(layer_key: str, delete_files: bool = False):
    """删除自定义矢量/栅格图层（内置图层 hx/caiqu 不允许删除）。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", layer_key):
        raise HTTPException(status_code=400, detail=f"非法的图层 key: {layer_key}")
    if layer_key in _TILE_LAYER_META:
        raise HTTPException(status_code=403, detail=f"内置图层不允许删除: {layer_key}")

    registry = _load_tile_registry()
    if layer_key not in registry:
        # 检查是否是无人机图层
        drone_registry = _load_drone_registry()
        if layer_key in drone_registry:
            meta = drone_registry.pop(layer_key)
            _save_drone_registry(drone_registry)
            return JSONResponse(content={"success": True, "layer": layer_key, "type": "drone", "deleted_files": []})

        # 图层不在任何注册表中 —— 尝试从 GeoServer 取消发布 + 可选删除 PostGIS 表
        gs_removed = False
        try:
            gs_result = _gs_client.unpublish_layer(layer_key, recurse=True)
            gs_removed = gs_result.get("removed", False)
        except _GeoServerUnavailable:
            pass
        except Exception as e:
            logger.warning("GeoServer unpublish failed for virtual layer %s: %s", layer_key, e)

        pg_dropped = False
        if delete_files:
            # 尝试删除 PostGIS 表
            try:
                conn = psycopg2.connect(**_PG_CONN)
                try:
                    conn.autocommit = True
                    with conn.cursor() as cur:
                        cur.execute(f'DROP TABLE IF EXISTS public."{layer_key}" CASCADE')
                        pg_dropped = True
                    logger.info("PostGIS table dropped: %s", layer_key)
                finally:
                    conn.close()
            except Exception as e:
                logger.warning("Drop PostGIS table failed for %s: %s", layer_key, e)

        # 即使 GeoServer 和 PostGIS 都没操作，也返回成功（前端会从列表中移除）
        logger.info("虚拟图层已删除: key=%s, gs_removed=%s, pg_dropped=%s", layer_key, gs_removed, pg_dropped)
        return JSONResponse(content={
            "success": True,
            "layer": layer_key,
            "type": "virtual",
            "gs_removed": gs_removed,
            "pg_dropped": pg_dropped,
            "deleted_files": [],
        })

    meta = registry.pop(layer_key)
    _save_tile_registry(registry)

    # 从内存栅格服务中注销
    _overlay_unregister_layer(layer_key)

    deleted_files: list = []
    if delete_files:
        # 删除矢量切片目录
        vt_dir = os.path.join(_VT_DIR, layer_key)
        if os.path.isdir(vt_dir):
            import shutil
            try:
                shutil.rmtree(vt_dir)
                deleted_files.append(vt_dir)
            except OSError as e:
                logger.warning("删除矢量切片目录失败: %s %s", vt_dir, e)
        # 删除源 GeoJSON（仅自定义图层）
        source_path = meta.get("source") or ""
        if source_path and os.path.isfile(source_path) and not source_path.startswith(_OVERLAY_DATA_DIR):
            try:
                os.remove(source_path)
                deleted_files.append(source_path)
            except OSError as e:
                logger.warning("删除源文件失败: %s %s", source_path, e)

    logger.info("图层已删除: key=%s, delete_files=%s, removed=%s", layer_key, delete_files, deleted_files)
    _invalidate_tile_stats()
    return JSONResponse(content={
        "success": True,
        "layer": layer_key,
        "deleted_files": deleted_files,
    })


@app.post("/api/tile_manager/build")
async def tile_manager_build(
    file: UploadFile = File(...),
    layer_key: str = Form(""),
    label: str = Form(""),
    build_type: str = Form("both"),
    stroke: str = Form("#2773d7"),
    fill: str = Form("#2773d7"),
    fill_alpha: float = Form(0.18),
    stroke_width: float = Form(2),
    point_size: float = Form(8),
    min_zoom: int = Form(0),
    max_zoom: int = Form(18),
    auto_publish: bool = Form(False),
):
    safe_key = _sanitize_layer_key(layer_key or os.path.splitext(file.filename or "")[0])
    if safe_key in _TILE_LAYER_META:
        raise HTTPException(status_code=400, detail="内置图层 key 不能覆盖")
    build_type = build_type if build_type in ("vector", "raster", "both") else "both"
    min_zoom = max(0, min(22, int(min_zoom)))
    max_zoom = max(min_zoom, min(22, int(max_zoom)))
    style = _parse_style(stroke, fill, fill_alpha, stroke_width, point_size)
    os.makedirs(_CUSTOM_TILE_DATA_DIR, exist_ok=True)
    source_path = os.path.join(_CUSTOM_TILE_DATA_DIR, f"{safe_key}.geojson")
    try:
        with open(source_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
        with open(source_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict) or data.get("type") not in ("FeatureCollection", "Feature"):
            raise ValueError("仅支持 GeoJSON FeatureCollection / Feature")
    except Exception as e:
        with contextlib.suppress(Exception):
            os.remove(source_path)
        raise HTTPException(status_code=400, detail=f"GeoJSON 文件无效: {e}")
    meta = {
        "label": label.strip() or safe_key,
        "color": style["stroke"],
        "source": source_path,
        "style": style,
        "min_zoom": min_zoom,
        "max_zoom": max_zoom,
        "build_type": build_type,
        "created_at": dt.now().isoformat(timespec="seconds"),
    }
    registry = _load_tile_registry()
    registry[safe_key] = meta
    _save_tile_registry(registry)
    if build_type in ("raster", "both"):
        _overlay_register_layer(safe_key, source_path, style)
    # 矢量切片生成与 GeoServer 发布放入后台任务，避免阻塞其他请求
    job_id = _submit_tile_build_job({}, lambda job: _run_geojson_build_job(
        job, safe_key, source_path, meta, build_type, min_zoom, max_zoom, style, auto_publish,
    ))
    return JSONResponse(content={
        "success": True,
        "layer": safe_key,
        "meta": meta,
        "async": True,
        "job_id": job_id,
    })


@app.get("/api/drone_imagery/layers")
async def drone_imagery_layers():
    layers = [
        _drone_layer_to_row(key, meta)
        for key, meta in _load_drone_registry().items()
        if isinstance(meta, dict)
    ]
    return JSONResponse(content={"success": True, "layers": layers})


@app.post("/api/drone_imagery/register")
async def drone_imagery_register(req: DroneImageryRegisterRequest):
    try:
        row = _register_drone_imagery(req)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return JSONResponse(content={"success": True, "layer": row["key"], "item": row})


@app.delete("/api/drone_imagery/{layer_key}")
async def drone_imagery_delete(layer_key: str, delete_files: bool = False):
    """删除无人机影像图层：从注册表中移除，并可选删除 MBTiles / 工作文件。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", layer_key):
        raise HTTPException(status_code=400, detail=f"非法的图层 key: {layer_key}")
    registry = _load_drone_registry()
    if layer_key not in registry:
        raise HTTPException(status_code=404, detail=f"无人机影像图层不存在: {layer_key}")
    meta = registry.pop(layer_key)
    _save_drone_registry(registry)

    deleted_files = []
    if delete_files and isinstance(meta, dict):
        # 删除 MBTiles 文件
        mbtiles_path = meta.get("path") or ""
        if mbtiles_path and os.path.isfile(mbtiles_path):
            try:
                os.remove(mbtiles_path)
                deleted_files.append(mbtiles_path)
            except OSError as e:
                logger.warning("删除 MBTiles 文件失败: %s %s", mbtiles_path, e)
        # 删除工作目录中的临时文件（如 _3857.tif）
        for suffix in ("_3857.tif", ".mbtiles"):
            work_file = os.path.join(_DRONE_WORK_DIR, f"{layer_key}{suffix}")
            if os.path.isfile(work_file):
                try:
                    os.remove(work_file)
                    deleted_files.append(work_file)
                except OSError as e:
                    logger.warning("删除工作文件失败: %s %s", work_file, e)

    logger.info("无人机影像图层已删除: key=%s, delete_files=%s, removed=%s", layer_key, delete_files, deleted_files)
    _invalidate_tile_stats()
    return JSONResponse(content={
        "success": True,
        "layer": layer_key,
        "deleted_files": deleted_files,
    })


@app.get("/api/file_browser")
async def file_browser(path: str = "/mnt", extensions: str = ".tif,.tiff"):
    """浏览服务器目录，返回子目录和文件列表（异步+超时保护）"""
    import pathlib, asyncio, concurrent.futures

    def _list_dir(dir_path: str, exts: str):
        target = pathlib.Path(dir_path).resolve()
        if not target.exists():
            return {"error": f"路径不存在: {dir_path}", "code": 404}
        if not target.is_dir():
            return {"error": f"不是目录: {dir_path}", "code": 400}
        ext_set = set(e.strip().lower() for e in exts.split(",") if e.strip())
        dirs_list, files_list = [], []
        try:
            for item in sorted(target.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                if item.name.startswith('.'):
                    continue
                if item.is_dir():
                    dirs_list.append({"name": item.name, "type": "dir"})
                elif item.is_file():
                    if ext_set and item.suffix.lower() not in ext_set:
                        continue
                    try:
                        size = item.stat().st_size
                    except OSError:
                        size = 0
                    files_list.append({"name": item.name, "type": "file", "size": size})
        except PermissionError:
            return {"error": f"无权访问: {dir_path}", "code": 403}
        parent = str(target.parent) if str(target) != "/" else None
        return {"path": str(target), "parent": parent, "dirs": dirs_list, "files": files_list}

    loop = asyncio.get_event_loop()
    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(None, _list_dir, path, extensions),
            timeout=10
        )
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail=f"读取目录超时（10秒），网络共享可能响应缓慢: {path}")
    if "error" in result:
        raise HTTPException(status_code=result["code"], detail=result["error"])
    return JSONResponse(content=result)


@app.post("/api/drone_imagery/build")
async def drone_imagery_build(req: DroneImageryBuildRequest):
    row = _run_drone_mbtiles_build(req)
    return JSONResponse(content={"success": True, "layer": row["key"], "item": row})


@app.post("/api/drone_imagery/build_stream")
async def drone_imagery_build_stream(req: DroneImageryBuildRequest):
    import asyncio, queue, threading

    def _producer(q: queue.Queue):
        try:
            for chunk in _run_drone_build_with_progress(req):
                q.put(chunk)
        except Exception as e:
            q.put(f"data: {json.dumps({'stage':'error','percent':0,'message':str(e)}, ensure_ascii=False)}\n\n")
        finally:
            q.put(None)

    async def _async_gen():
        q = queue.Queue()
        loop = asyncio.get_event_loop()
        loop.run_in_executor(None, _producer, q)
        while True:
            chunk = await loop.run_in_executor(None, q.get)
            if chunk is None:
                break
            yield chunk

    return StreamingResponse(_async_gen(), media_type="text/event-stream")


@app.post("/api/drone_imagery/build_async")
async def drone_imagery_build_async(req: DroneImageryBuildRequest):
    import threading
    job_id = uuid.uuid4().hex
    _DRONE_BUILD_JOBS[job_id] = {
        "job_id": job_id,
        "stage": "queued",
        "percent": 0,
        "message": "任务已提交，等待构建...",
        "done": False,
        "success": None,
        "created_at": dt.now().isoformat(timespec="seconds"),
        "updated_at": dt.now().isoformat(timespec="seconds"),
    }

    def _worker():
        try:
            chunks = _run_drone_build_with_progress(req)
            for chunk in chunks:
                if not chunk.startswith("data: "):
                    continue
                try:
                    payload = json.loads(chunk[6:].strip())
                except Exception:
                    continue
                job = _DRONE_BUILD_JOBS.get(job_id, {})
                job.update(payload)
                job["updated_at"] = dt.now().isoformat(timespec="seconds")
                if payload.get("stage") in ("done", "error"):
                    job["done"] = True
                    job["success"] = payload.get("stage") == "done"
                    job["finished_at"] = dt.now().isoformat(timespec="seconds")
                _DRONE_BUILD_JOBS[job_id] = job
        except Exception as e:
            payload = {"stage": "error", "percent": 0, "message": str(e)}
            job = _DRONE_BUILD_JOBS.get(job_id, {})
            job.update(payload)
            job["updated_at"] = dt.now().isoformat(timespec="seconds")
            job["done"] = True
            job["success"] = False
            job["finished_at"] = dt.now().isoformat(timespec="seconds")
            _DRONE_BUILD_JOBS[job_id] = job

    threading.Thread(target=_worker, daemon=True).start()
    return JSONResponse(content={"success": True, "job_id": job_id})


@app.get("/api/drone_imagery/build_status/{job_id}")
async def drone_imagery_build_status(job_id: str):
    job = _DRONE_BUILD_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="构建任务不存在或已过期")
    return JSONResponse(content={"success": True, "job": job})


@app.get("/api/drone_imagery/tile/{layer}/{z}/{x}/{y}.png")
async def drone_imagery_tile(layer: str, z: int, x: int, y: int):
    if not re.match(r"^[A-Za-z0-9_-]+$", layer):
        raise HTTPException(status_code=404, detail=f"unknown drone imagery layer: {layer}")
    if z < 0 or z > 22:
        raise HTTPException(status_code=400, detail="invalid zoom")
    n = 1 << z
    if x < 0 or x >= n or y < 0 or y >= n:
        raise HTTPException(status_code=400, detail="invalid tile xy")
    registry = _load_drone_registry()
    meta = registry.get(layer)
    if not isinstance(meta, dict):
        raise HTTPException(status_code=404, detail=f"unknown drone imagery layer: {layer}")
    path = os.path.abspath(meta.get("path") or "")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail=f"MBTiles 文件不存在: {path}")
    tile_row = (n - 1 - y) if meta.get("scheme", "tms") == "tms" else y
    try:
        with sqlite3.connect(path) as conn:
            row = conn.execute(
                "SELECT tile_data FROM tiles WHERE zoom_level=? AND tile_column=? AND tile_row=?",
                (z, x, tile_row),
            ).fetchone()
    except Exception as e:
        logger.exception("read drone imagery tile failed: %s/%s/%s/%s", layer, z, x, y)
        raise HTTPException(status_code=500, detail=str(e))
    if not row:
        return Response(status_code=204)
    media_type = _media_type_for_tile_format(str(meta.get("tile_format") or "png"))
    return Response(
        content=row[0],
        media_type=media_type,
        headers={
            "Cache-Control": "public, max-age=604800",
            "Access-Control-Allow-Origin": "*",
        },
    )


# ---------- 3D Tiles 管理 API ----------

@app.get("/api/tile_manager/3dtiles")
async def tile_manager_3dtiles_list():
    """列出所有已注册的 3D Tiles 数据集。"""
    registry = _load_3dtiles_registry()
    datasets = []
    for key, meta in registry.items():
        if isinstance(meta, dict):
            datasets.append(_3dtiles_layer_to_row(key, meta))
    return JSONResponse(content={"success": True, "datasets": datasets})


@app.post("/api/tile_manager/3dtiles/upload")
async def tile_manager_3dtiles_upload(
    file: UploadFile = File(...),
    key: str = Form(""),
    name: str = Form(""),
    label: str = Form(""),
    alt_offset: float = Form(0.0),
    auto_ground_clamp: bool = Form(True),
    description: str = Form(""),
):
    """上传 3D Tiles zip 包，自动解压并注册。"""
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="请选择 .zip 文件")
    safe_key = _sanitize_layer_key(key or os.path.splitext(file.filename)[0])
    target_dir = os.path.join(_3DTILES_DATA_DIR, safe_key)
    if os.path.exists(target_dir):
        raise HTTPException(status_code=409, detail=f"数据集 key 已存在: {safe_key}")
    os.makedirs(_3DTILES_DATA_DIR, exist_ok=True)
    tmp_path = ""
    try:
        import tempfile, zipfile
        with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp:
            _copy_upload_limited(file.file, tmp, _MAX_3DTILES_ZIP_BYTES)
            tmp_path = tmp.name
        os.makedirs(target_dir, exist_ok=True)
        _extract_zip_safely(tmp_path, target_dir)
        if tmp_path:
            os.unlink(tmp_path)
            tmp_path = ""
    except zipfile.BadZipFile:
        with contextlib.suppress(Exception):
            if tmp_path:
                os.unlink(tmp_path)
        shutil.rmtree(target_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail="无效的 zip 文件")
    except ValueError as e:
        with contextlib.suppress(Exception):
            if tmp_path:
                os.unlink(tmp_path)
        shutil.rmtree(target_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        with contextlib.suppress(Exception):
            if tmp_path:
                os.unlink(tmp_path)
        shutil.rmtree(target_dir, ignore_errors=True)
        raise
    except Exception as e:
        with contextlib.suppress(Exception):
            if tmp_path:
                os.unlink(tmp_path)
        shutil.rmtree(target_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"解压失败: {str(e)}")
    # 兼容嵌套目录：zip 内唯一子目录含 tileset.json 时，将内容提升到数据集根目录
    tileset_root = _locate_tileset_root(target_dir)
    if tileset_root is None:
        shutil.rmtree(target_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail="zip 包中未找到 tileset.json")
    if tileset_root != target_dir:
        try:
            for entry in os.listdir(tileset_root):
                src = os.path.join(tileset_root, entry)
                dst = os.path.join(target_dir, entry)
                if not os.path.exists(dst):
                    shutil.move(src, dst)
            shutil.rmtree(tileset_root, ignore_errors=True)
        except OSError as e:
            shutil.rmtree(target_dir, ignore_errors=True)
            raise HTTPException(status_code=500, detail=f"zip 目录结构处理失败: {e}")
    tileset_path = os.path.join(target_dir, "tileset.json")
    if not os.path.isfile(tileset_path):
        shutil.rmtree(target_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail="zip 包中未找到 tileset.json")
    meta_info = _read_3dtiles_meta(target_dir)
    _write_3dtiles_meta(target_dir, {
        "name": name.strip() or meta_info.get("name") or safe_key,
        "label": label.strip() or meta_info.get("label") or name.strip() or safe_key,
        "altOffset": float(alt_offset),
        "autoGroundClamp": bool(auto_ground_clamp),
        "description": description.strip() or meta_info.get("description") or "",
        "center": meta_info.get("center"),
    })
    stats = _count_3dtiles(target_dir)
    registry = _load_3dtiles_registry()
    registry[safe_key] = {
        "directory": target_dir,
        "label": label.strip() or name.strip() or safe_key,
        "name": name.strip() or safe_key,
        "tile_count": stats["tile_count"],
        "size_bytes": stats["size_bytes"],
        "alt_offset": float(alt_offset),
        "auto_ground_clamp": bool(auto_ground_clamp),
        "center": meta_info.get("center"),
        "description": description.strip() or "",
        "meta": meta_info,
        "created_at": dt.now().isoformat(timespec="seconds"),
    }
    _save_3dtiles_registry(registry)
    logger.info("3dtiles uploaded: key=%s, tiles=%s", safe_key, stats["tile_count"])
    return JSONResponse(content={
        "success": True,
        "layer": safe_key,
        "item": _3dtiles_layer_to_row(safe_key, registry[safe_key]),
    })


@app.post("/api/tile_manager/3dtiles/register")
async def tile_manager_3dtiles_register(req: ThreeDTilesRegisterRequest):
    """注册服务器上已有的 3D Tiles 目录。"""
    directory = os.path.abspath(req.directory)
    if not os.path.isdir(directory):
        raise HTTPException(status_code=404, detail=f"目录不存在: {directory}")
    tileset_path = os.path.join(directory, "tileset.json")
    if not os.path.isfile(tileset_path):
        raise HTTPException(status_code=400, detail="目录中未找到 tileset.json")
    safe_key = _sanitize_layer_key(req.key or os.path.basename(directory))
    registry = _load_3dtiles_registry()
    if safe_key in registry:
        raise HTTPException(status_code=409, detail=f"数据集 key 已存在: {safe_key}")
    meta_info = _read_3dtiles_meta(directory)
    _write_3dtiles_meta(directory, {
        "name": req.name.strip() or meta_info.get("name") or safe_key,
        "label": req.label.strip() or meta_info.get("label") or req.name.strip() or safe_key,
        "altOffset": float(req.alt_offset),
        "autoGroundClamp": bool(req.auto_ground_clamp),
        "description": req.description.strip() or meta_info.get("description") or "",
        "center": meta_info.get("center"),
    })
    stats = _count_3dtiles(directory)
    registry[safe_key] = {
        "directory": directory,
        "label": req.label.strip() or req.name.strip() or safe_key,
        "name": req.name.strip() or safe_key,
        "tile_count": stats["tile_count"],
        "size_bytes": stats["size_bytes"],
        "alt_offset": float(req.alt_offset),
        "auto_ground_clamp": bool(req.auto_ground_clamp),
        "center": meta_info.get("center"),
        "description": req.description.strip() or "",
        "meta": meta_info,
        "created_at": dt.now().isoformat(timespec="seconds"),
    }
    _save_3dtiles_registry(registry)
    logger.info("3dtiles registered: key=%s, directory=%s", safe_key, directory)
    return JSONResponse(content={
        "success": True,
        "layer": safe_key,
        "item": _3dtiles_layer_to_row(safe_key, registry[safe_key]),
    })


@app.post("/api/tile_manager/3dtiles/restats/{key}")
async def tile_manager_3dtiles_restats(key: str):
    """重新统计 3D Tiles 数据集的切片数与磁盘占用，并更新注册表缓存。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", key):
        raise HTTPException(status_code=400, detail=f"非法的 key: {key}")
    registry = _load_3dtiles_registry()
    if key not in registry:
        raise HTTPException(status_code=404, detail=f"数据集不存在: {key}")
    meta = registry[key]
    directory = meta.get("directory") or ""
    if not os.path.isdir(directory):
        raise HTTPException(status_code=404, detail=f"数据集目录不存在: {directory}")
    stats = _count_3dtiles(directory)
    meta["tile_count"] = stats["tile_count"]
    meta["size_bytes"] = stats["size_bytes"]
    meta["updated_at"] = dt.now().isoformat(timespec="seconds")
    _save_3dtiles_registry(registry)
    logger.info("3dtiles restats: key=%s, tiles=%s", key, stats["tile_count"])
    return JSONResponse(content={
        "success": True,
        "item": _3dtiles_layer_to_row(key, meta),
    })


@app.delete("/api/tile_manager/3dtiles/{key}")
async def tile_manager_3dtiles_delete(key: str, delete_files: bool = False):
    """删除已注册的 3D Tiles 数据集。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", key):
        raise HTTPException(status_code=400, detail=f"非法的 key: {key}")
    registry = _load_3dtiles_registry()
    if key not in registry:
        raise HTTPException(status_code=404, detail=f"数据集不存在: {key}")
    meta = registry.pop(key)
    _save_3dtiles_registry(registry)
    deleted_files = []
    if delete_files:
        directory = meta.get("directory") or ""
        if directory and os.path.isdir(directory):
            try:
                shutil.rmtree(directory)
                deleted_files.append(directory)
            except OSError as e:
                logger.warning("删除 3D Tiles 目录失败: %s %s", directory, e)
    logger.info("3dtiles deleted: key=%s, delete_files=%s, removed=%s", key, delete_files, deleted_files)
    _invalidate_tile_stats()
    return JSONResponse(content={
        "success": True,
        "layer": key,
        "deleted_files": deleted_files,
    })


@app.get("/api/3dtiles/{key}/tileset.json")
async def serve_3dtiles_tileset(key: str):
    """提供 3D Tiles 的 tileset.json 给 Cesium 加载。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", key):
        raise HTTPException(status_code=400, detail=f"非法的 key: {key}")
    registry = _load_3dtiles_registry()
    meta = registry.get(key)
    directory = (meta.get("directory") if isinstance(meta, dict) else None) or os.path.join(_3DTILES_DATA_DIR, key)
    tileset_path = os.path.join(directory, "tileset.json")
    if not os.path.isfile(tileset_path):
        raise HTTPException(status_code=404, detail="tileset.json 不存在")
    return FileResponse(
        tileset_path,
        media_type="application/json",
        headers={"Access-Control-Allow-Origin": "*", "Cache-Control": "public, max-age=3600"},
    )


@app.get("/api/3dtiles/{key}/{file_path:path}")
async def serve_3dtiles_file(key: str, file_path: str):
    """提供 3D Tiles 数据集内的任意文件（子瓦片、b3dm 等）给 Cesium 加载。"""
    if not re.match(r"^[A-Za-z0-9_-]+$", key):
        raise HTTPException(status_code=400, detail=f"非法的 key: {key}")
    # 防止路径穿越
    if ".." in file_path or file_path.startswith("/"):
        raise HTTPException(status_code=400, detail="非法的文件路径")
    registry = _load_3dtiles_registry()
    meta = registry.get(key)
    directory = (meta.get("directory") if isinstance(meta, dict) else None) or os.path.join(_3DTILES_DATA_DIR, key)
    full_path = os.path.join(directory, file_path)
    if not os.path.isfile(full_path):
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_path}")
    # 根据扩展名设置 MIME 类型
    ext = os.path.splitext(file_path)[1].lower()
    media_type_map = {
        ".json": "application/json",
        ".b3dm": "application/octet-stream",
        ".i3dm": "application/octet-stream",
        ".pnts": "application/octet-stream",
        ".cmpt": "application/octet-stream",
        ".gltf": "model/gltf+json",
        ".glb": "model/gltf-binary",
        ".bin": "application/octet-stream",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }
    media_type = media_type_map.get(ext, "application/octet-stream")
    return FileResponse(
        full_path,
        media_type=media_type,
        headers={"Access-Control-Allow-Origin": "*", "Cache-Control": "public, max-age=86400"},
    )


# ---------- 卫星影像动态切片（从本地 GeoTIFF 按需渲染） ----------
try:
    from tools.satellite_tile_service import get_satellite_tile, get_tile_bounds_4326, tiles_in_bounds
    _satellite_tile_available = True
    logger.info("satellite tile service loaded")
except Exception as _e:
    _satellite_tile_available = False
    logger.warning("satellite tile service NOT available: %s", _e)


@app.get("/api/satellite_tile/{z}/{x}/{y}.png")
async def satellite_tile(z: int, x: int, y: int):
    if not _satellite_tile_available:
        raise HTTPException(status_code=503, detail="satellite tile service not available")
    if z < 0 or z > 22:
        raise HTTPException(status_code=400, detail="invalid zoom")
    n = 1 << z
    if x < 0 or x >= n or y < 0 or y >= n:
        raise HTTPException(status_code=400, detail="invalid tile xy")
    png = get_satellite_tile(z, x, y)
    if png is None:
        return Response(status_code=204)
    return Response(
        content=png,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=3600", "Access-Control-Allow-Origin": "*"},
    )


# ==============================
# ✅ GeoServer 集成代理接口
# ==============================
class GeoServerPublishRequest(BaseModel):
    layer: str


class GeoServerSeedRequest(BaseModel):
    layer: str
    bounds: Optional[List[float]] = None  # [minx, miny, maxx, maxy] in EPSG:4326
    min_zoom: int = 0
    max_zoom: int = 14
    format: str = "image/png"
    threads: int = 1


def _find_tile_layer_meta(layer_key: str) -> Optional[Dict[str, Any]]:
    """从 /api/tile_manager/layers 同源逻辑里查 meta，避免重复请求。"""
    # 内置 + 自定义矢量/栅格
    merged = _merged_tile_meta()
    if layer_key in merged:
        meta = merged[layer_key]
        return {
            "type": "vector" if meta.get("build_type", "both") in ("vector", "both") else "raster",
            "label": meta.get("label", layer_key),
            "source_path": meta.get("source"),
            "style": meta.get("style") or {},
        }
    # 无人机
    drone = _load_drone_registry()
    if layer_key in drone:
        meta = drone[layer_key]
        # 优先使用 _3857.tif（保留时）；否则回退 source_path
        source_path = meta.get("source_path")
        warped = os.path.join(_DRONE_WORK_DIR, f"{layer_key}_3857.tif")
        if os.path.isfile(warped):
            source_path = warped
        return {
            "type": "drone",
            "label": meta.get("name", layer_key),
            "source_path": source_path,
        }
    if layer_key == "ceshen":
        return {"type": "vector", "label": "ceshen", "source_path": ""}
    return None


@app.get("/api/geoserver/status")
async def geoserver_status():
    try:
        return JSONResponse(content=_gs_client.health_status())
    except Exception as e:  # 兜底，避免任何意外影响主流程
        logger.warning("geoserver_status error: %s", e)
        return JSONResponse(content={"available": False, "reason": str(e)})


@app.get("/api/geoserver/layers")
async def geoserver_layers():
    try:
        items = _gs_client.list_layers()
        return JSONResponse(content={"available": True, "layers": items})
    except _GeoServerUnavailable as e:
        return JSONResponse(content={"available": False, "reason": str(e), "layers": []})
    except Exception as e:
        logger.warning("geoserver_layers error: %s", e)
        return JSONResponse(content={"available": False, "reason": str(e), "layers": []})


@app.get("/api/geoserver/capabilities")
async def geoserver_capabilities():
    return JSONResponse(content={
        "url": _gs_client.get_config()["url"],
        "workspace": _gs_client.get_config()["workspace"],
        "capabilities": _gs_client.capabilities_urls(),
    })


@app.get("/api/geoserver/preview/{layer}.png")
async def geoserver_preview(layer: str, bbox: Optional[str] = None, width: int = 520, height: int = 280):
    if not re.match(r"^[A-Za-z0-9_-]+$", layer):
        raise HTTPException(status_code=400, detail="invalid layer key")
    try:
        parsed_bbox = None
        if bbox:
            parts = [float(v) for v in bbox.split(",")]
            if len(parts) != 4:
                raise ValueError("bbox must contain 4 numbers")
            parsed_bbox = parts
        content, content_type = _gs_client.preview_image(layer, bbox=parsed_bbox, width=width, height=height)
        return Response(
            content=content,
            media_type=content_type,
            headers={"Cache-Control": "no-cache", "Access-Control-Allow-Origin": "*"},
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except _GeoServerUnavailable as e:
        raise HTTPException(status_code=502, detail=f"GeoServer 预览失败: {e}")


@app.post("/api/geoserver/publish")
async def geoserver_publish(req: GeoServerPublishRequest):
    if not re.match(r"^[A-Za-z0-9_-]+$", req.layer):
        raise HTTPException(status_code=400, detail="invalid layer key")
    meta = _find_tile_layer_meta(req.layer)
    if not meta:
        raise HTTPException(status_code=404, detail=f"unknown layer: {req.layer}")
    try:
        result = _gs_client.publish_by_tm_key(req.layer, meta)
        return JSONResponse(content={"success": True, **result})
    except _GeoServerUnavailable as e:
        raise HTTPException(status_code=502, detail=f"GeoServer 发布失败: {e}")


@app.post("/api/geoserver/unpublish")
async def geoserver_unpublish(req: GeoServerPublishRequest):
    if not re.match(r"^[A-Za-z0-9_-]+$", req.layer):
        raise HTTPException(status_code=400, detail="invalid layer key")
    try:
        result = _gs_client.unpublish_layer(req.layer, recurse=True)
        return JSONResponse(content={"success": True, "layer": req.layer, **result})
    except _GeoServerUnavailable as e:
        raise HTTPException(status_code=502, detail=f"GeoServer 取消发布失败: {e}")


@app.post("/api/geoserver/seed")
async def geoserver_seed(req: GeoServerSeedRequest):
    if not re.match(r"^[A-Za-z0-9_-]+$", req.layer):
        raise HTTPException(status_code=400, detail="invalid layer key")
    try:
        result = _gs_client.gwc_seed(
            req.layer,
            bounds=req.bounds,
            min_zoom=req.min_zoom,
            max_zoom=req.max_zoom,
            fmt=req.format,
            threads=req.threads,
        )
        return JSONResponse(content={"success": True, **result})
    except _GeoServerUnavailable as e:
        raise HTTPException(status_code=502, detail=f"GWC seed 失败: {e}")


@app.post("/api/geoserver/truncate")
async def geoserver_truncate(req: GeoServerSeedRequest):
    if not re.match(r"^[A-Za-z0-9_-]+$", req.layer):
        raise HTTPException(status_code=400, detail="invalid layer key")
    try:
        result = _gs_client.gwc_truncate(
            req.layer,
            bounds=req.bounds,
            min_zoom=req.min_zoom,
            max_zoom=req.max_zoom,
            fmt=req.format,
        )
        return JSONResponse(content={"success": True, **result})
    except _GeoServerUnavailable as e:
        raise HTTPException(status_code=502, detail=f"GWC truncate 失败: {e}")


@app.get("/api/geoserver/seed/{layer}")
async def geoserver_seed_status(layer: str):
    if not re.match(r"^[A-Za-z0-9_-]+$", layer):
        raise HTTPException(status_code=400, detail="invalid layer key")
    try:
        return JSONResponse(content=_gs_client.gwc_seed_status(layer))
    except _GeoServerUnavailable as e:
        raise HTTPException(status_code=502, detail=f"GWC 状态查询失败: {e}")


@app.get("/proxy/gf2023-tiles/{z}/{y}/{x}")
async def proxy_gf2023_tiles(z: int, y: int, x: int):
    """代理 2023年高分影像 GF_202308_cache 瓦片 (ArcGIS MapServer)"""
    url = f"http://123.149.20.94:60805/arcgis/rest/services/%E9%AB%98%E5%88%86%E5%BD%B1%E5%83%8F/GF_202308_cache/MapServer/tile/{z}/{y}/{x}"
    return await _proxy_arcgis_tile(url)


@app.get("/proxy/gf2026-tiles/{z}/{y}/{x}")
async def proxy_gf2026_tiles(z: int, y: int, x: int):
    """代理 2026年Q1 本地 GeoTIFF 高分影像瓦片 → serve_tile.py (port 8090)"""
    url = f"http://127.0.0.1:8090/tiles/{z}/{x}/{y}.png"
    return await _proxy_local_tile(url)


async def _proxy_local_tile(url: str):
    """通用本地瓦片代理（serve_tile.py）"""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(url)
        if r.status_code == 204:
            return Response(status_code=204)
        resp_headers = {
            "Access-Control-Allow-Origin": "*",
            "Cache-Control": "public, max-age=3600",
            "Content-Type": r.headers.get("Content-Type", "image/png"),
        }
        return Response(content=r.content, headers=resp_headers, status_code=r.status_code)
    except Exception as e:
        logger.error(f"代理本地瓦片失败: {url} - {e}")
        raise HTTPException(status_code=502, detail="本地瓦片服务不可达")


@app.get("/proxy/gf-tiles/{z}/{y}/{x}")
async def proxy_gf_tiles(z: int, y: int, x: int):
    """代理 GF_2024_YM 高分影像瓦片 (HTTP, /arcgis/ 路径)"""
    url = f"http://123.149.20.94:60805/arcgis/rest/services/%E9%AB%98%E5%88%86%E5%BD%B1%E5%83%8F/GF_2024_YM/MapServer/tile/{z}/{y}/{x}"
    return await _proxy_arcgis_tile(url)


@app.get("/proxy/gf2025-tiles/{z}/{y}/{x}")
async def proxy_gf2025_tiles(z: int, y: int, x: int):
    """代理 GF_202509_cache 高分影像瓦片"""
    url = f"http://123.149.20.94:60805/arcgis/rest/services/%E9%AB%98%E5%88%86%E5%BD%B1%E5%83%8F/GF_202509_cache/MapServer/tile/{z}/{y}/{x}"
    return await _proxy_arcgis_tile(url)


async def _proxy_arcgis_tile(url: str):
    """通用 ArcGIS 瓦片代理"""
    headers = {
        "User-Agent": "MapAssistant/1.0",
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    }
    try:
        async with httpx.AsyncClient(timeout=10, verify=False) as client:
            r = await client.get(url, headers=headers)
        content_type = r.headers.get("Content-Type", "image/png")
        resp_headers = {
            "Access-Control-Allow-Origin": "*",
            "Cache-Control": "public, max-age=600",
            "Content-Type": content_type,
        }
        return Response(content=r.content, headers=resp_headers, status_code=r.status_code)
    except Exception as e:
        logger.error(f"代理 ArcGIS 瓦片失败: {url} - {e}")
        raise HTTPException(status_code=502, detail="代理上游不可达")

@app.get("/suggestions")
async def get_suggestions():
    return {
        "suggestions": [
            "清除所有地图标记",
            "切换到卫星图层",
            "加载潢河郝楼可采区的矢量数据",
            "统计数据库中的点数量",
            "显示所有采样点并标注高程"
        ]
    }

# ==============================
# ✅ 新增：会话管理接口
# ==============================

class SessionCreate(BaseModel):
    title: Optional[str] = None

class SessionRename(BaseModel):
    title: str

@app.get("/api/sessions")
async def list_sessions():
    try:
        with contextlib.closing(get_db()) as conn:
            rows = conn.execute(
                "SELECT id, title, created_at, updated_at FROM sessions ORDER BY updated_at DESC"
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"List sessions error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/sessions")
async def create_session(body: SessionCreate):
    try:
        sid = str(uuid.uuid4())
        now = now_iso()
        title = (body.title or "新对话")[:50]
        with contextlib.closing(get_db()) as conn:
            conn.execute(
                "INSERT INTO sessions (id, title, created_at, updated_at) VALUES (?,?,?,?)",
                (sid, title, now, now)
            )
            conn.commit()
        return {"id": sid, "title": title, "created_at": now, "updated_at": now}
    except Exception as e:
        logger.error(f"Create session error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/sessions/{session_id}")
async def delete_session(session_id: str):
    try:
        with contextlib.closing(get_db()) as conn:
            conn.execute("PRAGMA foreign_keys=ON")
            conn.execute("DELETE FROM sessions WHERE id=?", (session_id,))
            conn.commit()
        return {"success": True}
    except Exception as e:
        logger.error(f"Delete session error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/sessions/{session_id}")
async def rename_session(session_id: str, body: SessionRename):
    try:
        title = body.title[:50]
        with contextlib.closing(get_db()) as conn:
            conn.execute(
                "UPDATE sessions SET title=?, updated_at=? WHERE id=?",
                (title, now_iso(), session_id)
            )
            conn.commit()
        return {"success": True, "title": title}
    except Exception as e:
        logger.error(f"Rename session error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/sessions/{session_id}/messages")
async def get_session_messages(session_id: str):
    try:
        with contextlib.closing(get_db()) as conn:
            rows = conn.execute(
                "SELECT role, content, created_at FROM messages WHERE session_id=? ORDER BY id ASC",
                (session_id,)
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"Get session messages error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==============================
# 知识库管理接口 (RagFlow)
# ==============================
@app.get("/api/knowledge")
async def list_knowledge(project: str = None):
    try:
        kb_tool = KnowledgeBaseTool()
        req_params = {'operation': 'list_topics'}
        if project:
            req_params['project'] = project
        result = kb_tool.call(req_params)
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error'))
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"List knowledge error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/knowledge/folders")
async def list_knowledge_folders():
    """列出知识库项目文件夹及各自文档数（供前端“选择文件夹”入口）。"""
    try:
        kb_tool = KnowledgeBaseTool()
        result = kb_tool.call({'operation': 'list_folders'})
        # 兼容不支持文件夹的后端（如 RagFlow）：返回空列表而非报错
        if not result.get('success'):
            return {'success': True, 'folders': [], 'total_folders': 0, 'total_documents': 0}
        return result
    except Exception as e:
        logger.error(f"List knowledge folders error: {e}")
        return {'success': True, 'folders': [], 'total_folders': 0, 'total_documents': 0}


@app.get("/api/knowledge/graph")
async def get_knowledge_graph():
    """知识图谱可视化数据接口 — 返回节点和关系用于前端力导向图渲染"""
    try:
        kg = get_kg()
        loop = asyncio.get_event_loop()
        data = await loop.run_in_executor(None, kg.get_graph_data)
        return {"success": True, "data": data}
    except Exception as e:
        logger.error(f"Get knowledge graph error: {e}")
        return {"success": False, "error": str(e)}

@app.get("/api/knowledge/{document_id}")
async def get_knowledge_content(document_id: str):
    try:
        kb_tool = KnowledgeBaseTool()
        result = kb_tool.call({
            'operation': 'get_content',
            'document_id': document_id
        })
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error'))
        return result
    except Exception as e:
        logger.error(f"Get knowledge content error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/knowledge/qa")
async def knowledge_qa(req: KnowledgeQARequest):
    """
    智能问答：KnowledgeQAAgent 四阶段管道
    Stage 1: 查询理解（Qwen 提取实体/属性/时间）
    Stage 2: 多路检索（RagFlow 原问题+关键词+宽泛查询）
    Stage 3: 结构化提取（Qwen 从 chunk 提取 JSON）
    Stage 4: 推理生成（Qwen 合成最终答案 + 来源引用）
    """
    try:
        agent = KnowledgeQAAgent()
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, agent.answer, req.question, req.top_k)
        return result
    except Exception as e:
        logger.error(f"Knowledge QA error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/knowledge/add")
async def add_knowledge(req: KnowledgeAddRequest):
    """添加文本文档到 RagFlow"""
    try:
        kb_tool = KnowledgeBaseTool()
        result = kb_tool.call({
            'operation': 'add_document',
            'name': req.name,
            'content': req.content
        })
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error'))
        return result
    except Exception as e:
        logger.error(f"Add knowledge error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/knowledge/upload")
async def upload_knowledge_file(file: UploadFile = File(...)):
    """上传文件到 RagFlow"""
    import tempfile as _tmp
    tmp_path = None
    try:
        # 保存上传文件到临时目录
        suffix = os.path.splitext(file.filename or "document.txt")[1] or ".txt"
        with _tmp.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        kb_tool = KnowledgeBaseTool()
        result = kb_tool.call({
            'operation': 'upload_file',
            'file_path': tmp_path
        })
        
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error'))
        return result
    except Exception as e:
        logger.error(f"Upload knowledge file error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

@app.get("/api/knowledge/diagnose/stats")
async def diagnose_kb_stats():
    """知识库索引诊断统计 - 返回 chunk 数、文档数、存储大小等"""
    try:
        kb_tool = KnowledgeBaseTool()
        if not kb_tool._ensure_initialized():
            return {"success": False, "error": "LlamaIndex 未初始化"}
        
        store = kb_tool._index.docstore
        all_docs = list(store.docs.items())
        real_docs = [(k, v) for k, v in all_docs if k != "placeholder"]
        
        # 按文档分组
        doc_groups = {}
        for doc_id, doc in real_docs:
            meta = doc.metadata or {}
            parent_id = meta.get("document_id", doc_id)
            doc_groups.setdefault(parent_id, []).append((doc_id, doc))
        
        # 详细文档信息
        doc_list = []
        for parent_id, chunks in doc_groups.items():
            first_chunk = chunks[0][1] if chunks else None
            name = first_chunk.metadata.get("title", parent_id[:40]) if first_chunk else parent_id[:40]
            total_chars = sum(len(c.text) for _, c in chunks)
            avg_embedding = sum(len(c.embedding) if c.embedding else 0 for _, c in chunks) / max(len(chunks), 1)
            doc_list.append({
                "id": parent_id,
                "name": name,
                "chunk_count": len(chunks),
                "total_chars": total_chars,
                "avg_embedding_dim": round(avg_embedding),
                "created_at": first_chunk.metadata.get("created_at", "") if first_chunk else ""
            })
        
        doc_list.sort(key=lambda x: x["chunk_count"], reverse=True)
        
        # 存储文件大小
        import glob
        persist_dir = kb_tool._persist_dir
        file_sizes = {}
        for f in glob.glob(os.path.join(persist_dir, "*.json")):
            fname = os.path.basename(f)
            file_sizes[fname] = os.path.getsize(f)
        
        return {
            "success": True,
            "total_chunks": len(real_docs),
            "total_documents": len(doc_groups),
            "persist_dir": persist_dir,
            "chunk_size": kb_tool._index.settings.chunk_size if hasattr(kb_tool._index, 'settings') else 512,
            "chunk_overlap": kb_tool._index.settings.chunk_overlap if hasattr(kb_tool._index, 'settings') else 64,
            "embed_model": kb_tool._embed_model_name,
            "file_sizes": file_sizes,
            "documents": doc_list[:50]  # 最多返回 50 个
        }
    except Exception as e:
        logger.error(f"Diagnose stats error: {e}")
        return {"success": False, "error": str(e)}


@app.delete("/api/knowledge/{kb_id}")
async def delete_knowledge(kb_id: str):
    """从 RagFlow 删除文档"""
    try:
        kb_tool = KnowledgeBaseTool()
        result = kb_tool.call({
            'operation': 'delete_document',
            'document_id': kb_id
        })
        if not result.get('success'):
            raise HTTPException(status_code=500, detail=result.get('error'))
        return result
    except Exception as e:
        logger.error(f"Delete knowledge error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/memory/facts")
async def list_user_facts(q: str = ""):
    """查看用户事实记忆（阶段D，支持 q 关键字过滤）。"""
    from agents.fact_memory import list_facts
    try:
        facts = list_facts(q=q or None)
        return {"success": True, "total": len(facts), "facts": facts}
    except Exception as e:
        logger.error(f"List facts error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/memory/facts/{fact_id}")
async def delete_user_fact(fact_id: str):
    """删除指定用户事实记忆。"""
    from agents.fact_memory import delete_fact
    try:
        if not delete_fact(fact_id):
            raise HTTPException(status_code=404, detail="事实不存在")
        return {"success": True, "id": fact_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete fact error: {e}")
        raise HTTPException(status_code=500, detail=str(e))



def _schedule_fact_extraction(session_id: str, user_text: str, assistant_text: str):
    """run 成功收尾后 fire-and-forget 抽取用户事实记忆（阶段D）。同步/异步上下文通吃。"""
    try:
        from agents.fact_memory import extract_and_store_async
        llm = getattr(task_executor, "_llm", None) if task_executor is not None else None
        extract_and_store_async(session_id, user_text, assistant_text, llm)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[fact-memory] 调度失败: {e}")


def _persist_chat(session_id: str, user_content: str, assistant_content: str):
    """将本轮对话持久化到 SQLite，如果 session_id 不存在则跳过。"""
    if not session_id or session_id == "default":
        return
    try:
        with contextlib.closing(get_db()) as conn:
            row = conn.execute("SELECT id FROM sessions WHERE id=?", (session_id,)).fetchone()
            if not row:
                return  # 不是有效会话，跳过
            now = now_iso()
            conn.execute(
                "INSERT INTO messages (session_id, role, content, created_at) VALUES (?,?,?,?)",
                (session_id, "user", user_content, now)
            )
            conn.execute(
                "INSERT INTO messages (session_id, role, content, created_at) VALUES (?,?,?,?)",
                (session_id, "assistant", assistant_content, now)
            )
            # 自动命名：如果是第一条消息，用前20字作为标题
            msg_count = conn.execute(
                "SELECT COUNT(*) FROM messages WHERE session_id=?", (session_id,)
            ).fetchone()[0]
            if msg_count <= 2:  # 刚插入的这两条
                title = user_content[:20] + ("…" if len(user_content) > 20 else "")
                conn.execute(
                    "UPDATE sessions SET title=?, updated_at=? WHERE id=?",
                    (title, now, session_id)
                )
            else:
                conn.execute(
                    "UPDATE sessions SET updated_at=? WHERE id=?",
                    (now, session_id)
                )
            conn.commit()
    except Exception as e:
        logger.warning(f"_persist_chat failed (non-fatal): {e}")


def clean_response_content(content: str) -> str:
    if not content:
        return content
    content = re.sub(r'<\w+_tool[^>]*>', '', content)
    content = re.sub(r'\n\s*\n+', '\n\n', content)
    return content.strip()


def _is_explicit_map_request(user_content: str) -> bool:
    if not user_content:
        return False
    map_terms = [
        "地图", "上图", "加载", "图层", "定位", "跳转", "飞到", "显示到地图",
        "加载到地图", "标记", "标注", "打点", "落点", "经纬度", "卫星图", "底图",
        "切换", "卫星", "清除",
        # qgis_mcp_tool 产出结果也需要地图展示
        "中心点", "缓冲区", "buffer", "裁剪", "clip",
        # 空间分析类结果也需要地图展示
        "距离", "连线", "最近", "最短", "多远",
    ]
    return any(term in user_content for term in map_terms)


def _is_explicit_marker_request(user_content: str) -> bool:
    if not user_content:
        return False
    marker_terms = ["标记", "标注", "打点", "落点", "经纬度", "坐标点"]
    return any(term in user_content for term in marker_terms)


def _stable_json_key(payload: Any) -> str:
    try:
        return json.dumps(payload, ensure_ascii=False, sort_keys=True)
    except Exception:
        return str(payload)


def _optimize_map_commands(user_content: str, map_commands: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not map_commands:
        return []

    if not _is_explicit_map_request(user_content):
        logger.info("Suppressing map commands for non-map visualization request.")
        return []

    explicit_marker = _is_explicit_marker_request(user_content)
    seen = set()
    unique_commands = []
    for cmd in map_commands:
        key = _stable_json_key(cmd)
        if key in seen:
            continue
        seen.add(key)
        unique_commands.append(cmd)

    switch_layer_cmd = None
    set_view_cmd = None
    vector_commands = []
    marker_commands = []
    fit_markers_cmd = None
    other_commands = []

    for cmd in unique_commands:
        cmd_type = cmd.get("type")
        if cmd_type == "switch_layer":
            switch_layer_cmd = cmd
        elif cmd_type == "set_view" and set_view_cmd is None:
            set_view_cmd = cmd
        elif cmd_type == "load_vector_layer":
            vector_commands.append(cmd)
        elif cmd_type == "add_marker":
            if explicit_marker and len(marker_commands) < 3:
                marker_commands.append(cmd)
        elif cmd_type == "fit_markers":
            fit_markers_cmd = cmd
        else:
            other_commands.append(cmd)

    optimized = []
    if switch_layer_cmd:
        optimized.append(switch_layer_cmd)

    if vector_commands:
        optimized.extend(vector_commands[:2])
    else:
        if set_view_cmd:
            optimized.append(set_view_cmd)
        optimized.extend(marker_commands)
        if fit_markers_cmd and marker_commands:
            optimized.append(fit_markers_cmd)

    optimized.extend(other_commands)

    logger.info(
        "Optimized map commands from %s to %s (explicit_marker=%s)",
        len(map_commands), len(optimized), explicit_marker
    )
    return optimized


def _optimize_charts(charts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not charts:
        return []

    unique = []
    seen = set()
    for chart in charts:
        config = chart.get("config") or {}
        title = ((config.get("title") or {}).get("text") if isinstance(config.get("title"), dict) else None) or chart.get("summary") or chart.get("chart_type")
        key = f"{chart.get('chart_type', 'chart')}::{title}::{_stable_json_key(config)}"
        if key in seen:
            continue
        seen.add(key)
        unique.append(chart)

    if len(unique) <= 2:
        return unique

    selected = []
    used_types = set()
    for chart in unique:
        chart_type = chart.get("chart_type", "chart")
        if chart_type not in used_types:
            selected.append(chart)
            used_types.add(chart_type)
        if len(selected) >= 2:
            break

    if not selected:
        selected = unique[:2]

    logger.info("Optimized charts from %s to %s", len(charts), len(selected))
    return selected

@app.get("/api/mvt/{z}/{x}/{y}")
async def get_mvt_tile(
    z: int,
    x: int,
    y: int,
    table_name: str = "ceshen",
    geom_col: str = "geom",
    properties: str = None,
    filter: str = None
):
    try:
        if not re.match(r'^[a-zA-Z0-9_"\u4e00-\u9fa5\s]+$', table_name):
            raise HTTPException(status_code=400, detail="无效的表名格式")
        if not re.match(r'^[a-zA-Z0-9_"\u4e00-\u9fa5\s]+$', geom_col):
            raise HTTPException(status_code=400, detail="无效的几何列格式")

        n = 2 ** z
        lon_min = x / n * 360.0 - 180.0
        lon_max = (x + 1) / n * 360.0 - 180.0
        import math
        def tile2lat(ty, tz):
            n_ = math.pi - 2.0 * math.pi * ty / (2.0 ** tz)
            return math.degrees(math.atan(math.sinh(n_)))
        lat_max = tile2lat(y, z)
        lat_min = tile2lat(y + 1, z)

        props_select = ""
        if properties:
            fields = [p.strip() for p in properties.split(",") if p.strip()]
            cleaned = []
            for f in fields:
                if f.startswith('"') and f.endswith('"'):
                    cleaned.append(f)
                else:
                    cleaned.append(f'"{f}"')
            for f in cleaned:
                props_select += f", t.{f}"

        safe_table = table_name if table_name.startswith('"') else f'"{table_name}"'
        env_sql = f"ST_Transform(ST_MakeEnvelope({lon_min}, {lat_min}, {lon_max}, {lat_max}, 4326), 3857)"
        where_extra = ""
        if filter:
            where_extra = f" AND ({filter})"

        sql = f"""
        SELECT encode(ST_AsMVT(q, '{table_name}', 4096, 'geom'), 'base64') AS tile
        FROM (
            SELECT
                ST_AsMVTGeom(
                    ST_Transform(t.{geom_col}, 3857),
                    b.env,
                    4096,
                    256,
                    true
                ) AS geom
                {props_select}
            FROM {safe_table} t
            JOIN (SELECT {env_sql} AS env) b ON TRUE
            WHERE t.{geom_col} IS NOT NULL
              AND ST_IsValid(t.{geom_col})
              AND ST_Intersects(ST_Transform(t.{geom_col}, 3857), b.env)
              {where_extra}
        ) AS q;
        """

        conn = psycopg2.connect(**_PG_CONN)
        try:
            with conn.cursor() as cur:
                cur.execute(sql)
                row = cur.fetchone()
                if not row or not row[0]:
                    return Response(content=b"", media_type="application/x-protobuf", headers={"Cache-Control": "public, max-age=300"})
                import base64
                tile_bytes = base64.b64decode(row[0])
                return Response(content=tile_bytes, media_type="application/x-protobuf", headers={"Cache-Control": "public, max-age=300"})
        finally:
            try:
                conn.close()
            except:
                pass
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"MVT error: {e}")
        raise HTTPException(status_code=500, detail="生成矢量瓦片失败")


# ==============================
# SAM 目标识别接口
# ==============================
class SAMDetectRequest(BaseModel):
    geometry: dict  # GeoJSON Polygon geometry
    prompt: str
    mode: str = "rectangle"  # rectangle | polygon
    fast_mode: bool = False
    quick_mode: bool = False
    precise_mode: bool = False  # 精度模式：默认关闭（快速优先）


@app.post("/api/sam-detect")
async def sam_detect(req: SAMDetectRequest):
    """
    SAM 目标识别：根据绘制的 GeoJSON 区域和文本提示词，执行 SAM 推理
    通过 SSE 实时推送进度，最终返回 GeoJSON 结果
    """
    import subprocess
    import tempfile
    import shutil
    import uuid
    import asyncio
    import queue
    import threading

    logger.info(f"SAM detect request: prompt={req.prompt}, mode={req.mode}, precise_mode={req.precise_mode}")

    # 预检：坐标是否在影像覆盖范围内
    coords = req.geometry.get("coordinates", [])
    if coords:
        flat = coords[0] if isinstance(coords[0][0], (int, float)) else coords[0][0] if coords else []
        if isinstance(flat[0], (int, float)):
            def extract_points(c):
                if isinstance(c[0], (int, float)):
                    return [c]
                return [p for sub in c for p in (extract_points(sub) if isinstance(sub[0], list) else [sub])]
            points = extract_points(coords)
            lons = [p[0] for p in points]
            lats = [p[1] for p in points]
            tif_bounds = (110.35, 116.65, 31.38, 36.37)
            if max(lons) < tif_bounds[0] or min(lons) > tif_bounds[1] or \
               max(lats) < tif_bounds[2] or min(lats) > tif_bounds[3]:
                raise HTTPException(
                    status_code=400,
                    detail=f"绘制区域超出影像覆盖范围。影像覆盖: 经度 110.35~116.65, 纬度 31.38~36.37"
                )

    task_id = uuid.uuid4().hex[:12]
    progress_dir = "/tmp/sam_progress"
    os.makedirs(progress_dir, exist_ok=True)
    progress_file = os.path.join(progress_dir, f"{task_id}.json")

    # 初始化进度文件
    with open(progress_file, 'w') as f:
        json.dump({"stage": "init", "current": 0, "total": 1, "message": "任务已创建，正在启动..."}, f)

    output_dir = tempfile.mkdtemp(prefix="sam_detect_")
    sam_script = "/home/server/python/map_assistant_v1/backend/tools/sam_predict.py"
    python_bin = "/home/server/miniconda3/envs/sam/bin/python"
    geometry_json = json.dumps(req.geometry)
    cmd = [python_bin, "-u", sam_script, geometry_json, req.prompt]
    if not req.precise_mode:
        cmd.append("--demo")
    if req.quick_mode:
        cmd.append("--quick")

    logger.info(f"SAM command (streaming): {' '.join(cmd)}")

    # ── SSE 流式生成器 ──
    def _run_and_stream(q: queue.Queue):
        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                env={**os.environ,
                     "CUDA_VISIBLE_DEVICES": os.environ.get("CUDA_VISIBLE_DEVICES", "0"),
                     "SAM_PROGRESS_FILE": progress_file,
                     }
            )

            # 逐行读取 stdout，解析进度和结果
            stdout_lines = []
            for line in proc.stdout:
                line = line.rstrip('\n').rstrip('\r')
                if not line:
                    continue
                stdout_lines.append(line)

                # 解析 [PROGRESS] 行 → SSE 进度事件
                if line.startswith("[PROGRESS]"):
                    try:
                        parts = line[10:].strip().split(" ", 2)
                        stage = parts[0] if len(parts) > 0 else ""
                        current_total = parts[1].split("/") if len(parts) > 1 else ["0", "1"]
                        current = int(current_total[0]) if current_total[0].isdigit() else 0
                        total = int(current_total[1]) if len(current_total) > 1 and current_total[1].isdigit() else max(current, 1)
                        message = parts[2] if len(parts) > 2 else ""

                        # 阶段 → 百分比映射
                        pct_map = {
                            "init": 5, "cropped": 12, "inference_start": 18,
                        }
                        if stage in pct_map:
                            pct = pct_map[stage]
                        elif stage == "inference":
                            pct = min(18 + int((current / max(total, 1)) * 72), 90)
                        elif stage == "inference_done":
                            pct = 92
                        elif stage == "done":
                            pct = 100
                        elif stage == "error":
                            pct = 0
                        else:
                            pct = 50  # fallback

                        q.put(json.dumps({
                            "type": "progress", "stage": stage,
                            "percent": pct, "message": message,
                        }, ensure_ascii=False))
                    except Exception:
                        pass
                elif line.startswith("[SAM]"):
                    # 日志行也推送为进度消息
                    q.put(json.dumps({
                        "type": "progress", "stage": "log",
                        "percent": -1, "message": line,
                    }, ensure_ascii=False))

            proc.wait()

            # 读取 stderr
            stderr_text = proc.stderr.read()

            if proc.returncode != 0:
                err_msg = stderr_text[:500] if stderr_text else f"进程退出码 {proc.returncode}"
                q.put(json.dumps({
                    "type": "error", "message": err_msg,
                }, ensure_ascii=False))
                return

            # 在 stdout_lines 中找最后一行 JSON（feature collection）
            result_data = None
            for line in reversed(stdout_lines):
                line = line.strip()
                if line.startswith("{"):
                    try:
                        result_data = json.loads(line)
                        if isinstance(result_data, dict) and "features" in result_data:
                            break
                        result_data = None
                    except json.JSONDecodeError:
                        continue

            if result_data is None:
                # 再试 stderr
                if "超出影像范围" in stderr_text:
                    q.put(json.dumps({
                        "type": "error", "message": "绘制区域超出影像覆盖范围",
                    }, ensure_ascii=False))
                else:
                    q.put(json.dumps({
                        "type": "error", "message": "SAM 输出解析失败",
                    }, ensure_ascii=False))
                return

            result_data["_task_id"] = task_id
            logger.info(f"SAM result: {len(result_data.get('features', []))} features")
            q.put(json.dumps({
                "type": "final", "result": result_data,
            }, ensure_ascii=False))

        except Exception as e:
            q.put(json.dumps({
                "type": "error", "message": str(e),
            }, ensure_ascii=False))
        finally:
            q.put(None)  # 结束信号

    async def _sse_generator():
        q = queue.Queue()
        loop = asyncio.get_event_loop()
        loop.run_in_executor(None, _run_and_stream, q)

        while True:
            chunk = await loop.run_in_executor(None, q.get)
            if chunk is None:
                break
            yield f"data: {chunk}\n\n"

    return StreamingResponse(
        _sse_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ==============================
# SAM 变化检测接口 — 已移除（统一为 SAM 语义分割路径）
# 原 /api/sam-change-detect 端点已删除，变化检测功能不再支持
# ==============================


# ==============================
# SAM 阈值配置接口 — 已移除（VLM 相关功能已删除）
# 统一使用 SAM3 语义分割，不再需要 VLM 阈值调节
# ==============================
# 瓦片级 VLM 检测（Gemma4 六步流程）— 已移除
# 原 /api/sam-tile-detect 端点及相关函数已删除
# 检测路径统一为 SAM3 语义分割
# ==============================


@app.get("/api/sam-progress/{task_id}")
async def sam_progress(task_id: str):
    """查询 SAM 推理实时进度"""
    progress_file = f"/tmp/sam_progress/{task_id}.json"
    if not os.path.exists(progress_file):
        raise HTTPException(status_code=404, detail="任务不存在或已过期")
    try:
        with open(progress_file) as f:
            return json.load(f)
    except Exception as e:
        return {"stage": "error", "current": 0, "total": 1, "message": str(e)}


@app.get("/api/sam-test-result")
async def sam_test_result():
    """返回预计算的 SAM 测试结果（南阳市区建筑检测，365个图斑）"""
    test_file = "/tmp/sam_building_result.json"
    if os.path.exists(test_file):
        with open(test_file) as f:
            data = json.load(f)
        data["_note"] = "测试数据：南阳市区 (112.52~112.54, 33.01~33.03) 建筑检测结果"
        return data
    return {"type": "FeatureCollection", "features": [], "_note": "测试数据尚未生成"}


class SAMDownloadRequest(BaseModel):
    geojson: dict


@app.post("/api/sam-download")
async def sam_download(req: SAMDownloadRequest):
    """
    将 SAM 识别结果 GeoJSON 打包为 SHP + ZIP 下载
    """
    import zipfile
    import tempfile
    import io

    geojson = req.geojson
    if not geojson or not geojson.get("features"):
        raise HTTPException(status_code=400, detail="无识别结果可下载")

    try:
        import geopandas as gpd

        # 转 GeoDataFrame
        gdf = gpd.GeoDataFrame.from_features(geojson["features"], crs="EPSG:4326")

        # 写入临时目录
        tmp_dir = tempfile.mkdtemp(prefix="sam_shp_")
        shp_path = os.path.join(tmp_dir, "sam_result.shp")
        gdf.to_file(shp_path)

        # 打包 ZIP
        zip_path = os.path.join(tmp_dir, "sam_result.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                fpath = shp_path.replace('.shp', ext)
                if os.path.exists(fpath):
                    zf.write(fpath, os.path.basename(fpath))

        # 读取 ZIP 返回
        with open(zip_path, 'rb') as f:
            content = f.read()

        # 清理
        shutil.rmtree(tmp_dir, ignore_errors=True)

        return Response(
            content=content,
            media_type="application/zip",
            headers={"Content-Disposition": 'attachment; filename="sam_result.zip"'}
        )
    except Exception as e:
        logger.exception(f"SHP download error: {e}")
        raise HTTPException(status_code=500, detail=f"打包下载失败: {str(e)}")


# ==============================
# SAM Requery 精修 API
# ==============================

class SAMRequeryRequest(BaseModel):
    geometry: dict           # GeoJSON Polygon — 原始裁剪区域
    prompt: str              # 文本提示词
    coarse_geojson: dict     # 粗检测 GeoJSON FeatureCollection（/api/sam-detect 的输出）


@app.post("/api/sam-requery")
async def sam_requery(req: SAMRequeryRequest):
    """
    Requery 精修：对粗检测结果中的每个图斑，裁剪对应影像区域 →
    SAM3 再推理 → 返回精修 GeoJSON。

    输入：
      - geometry: 原始绘制区域（用于从 TIF 裁剪影像）
      - prompt: 文本提示词
      - coarse_geojson: 粗检测 GeoJSON FeatureCollection
    返回：
      精修后的 GeoJSON FeatureCollection
    """
    import subprocess
    import tempfile
    import shutil
    import uuid

    logger.info(f"SAM Requery: prompt={req.prompt[:30]}, coarse_features={len(req.coarse_geojson.get('features', []))}")

    task_id = uuid.uuid4().hex[:12]
    output_dir = tempfile.mkdtemp(prefix="sam_requery_")

    try:
        sam_script = "/home/server/python/map_assistant_v1/backend/tools/sam_predict.py"
        python_bin = "/home/server/miniconda3/envs/sam/bin/python"

        geometry_json = json.dumps(req.geometry)
        coarse_json = json.dumps(req.coarse_geojson)

        cmd = [python_bin, "-u", sam_script, geometry_json, req.prompt, "--requery"]

        logger.info(f"SAM Requery command: {' '.join(cmd)}")
        proc = subprocess.run(
            cmd,
            input=coarse_json,
            capture_output=True,
            text=True,
            timeout=600,
            env={**os.environ,
                 "CUDA_VISIBLE_DEVICES": os.environ.get("CUDA_VISIBLE_DEVICES", "0"),
                 }
        )

        if proc.returncode != 0:
            logger.error(f"SAM Requery failed: {proc.stderr}")
            raise HTTPException(status_code=500, detail=f"Requery 推理失败: {proc.stderr[:500]}")

        # 解析 JSON 输出（取最后一行）
        output_lines = proc.stdout.strip().split("\n")
        result_json = None
        for line in reversed(output_lines):
            line = line.strip()
            if line.startswith("{"):
                try:
                    result_json = json.loads(line)
                    break
                except json.JSONDecodeError:
                    continue

        if result_json is None:
            logger.error(f"Requery output parse failed: {proc.stdout[-500:]}")
            raise HTTPException(status_code=500, detail="Requery 输出解析失败")

        n_refined = sum(1 for f in result_json.get("features", [])
                        if f.get("properties", {}).get("refined"))
        logger.info(f"SAM Requery done: {len(result_json.get('features', []))} features, {n_refined} refined")
        result_json["_task_id"] = task_id
        return result_json

    except subprocess.TimeoutExpired:
        logger.error("SAM Requery timeout")
        raise HTTPException(status_code=504, detail="Requery 推理超时")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"SAM Requery error: {e}")
        raise HTTPException(status_code=500, detail=f"Requery 异常: {str(e)}")
    finally:
        try:
            shutil.rmtree(output_dir, ignore_errors=True)
        except Exception:
            pass


# ==============================
# 标注模块 API（交互式遥感影像标注）
# ==============================

from tools.annotation_store import (
    save as annotation_save,
    save_batch,
    load_by_session,
    load_by_id as annotation_load_by_id,
    delete as annotation_delete,
    delete_by_session,
    list_sessions,
    export_geojson as annotation_export_geojson,
    export_coco as annotation_export_coco,
)


class AnnotationItem(BaseModel):
    id: str | None = None
    session_id: str = ""
    image_path: str = ""
    label: str = ""
    class_id: int | None = None
    geometry: dict = {}
    mask_path: str | None = None
    source: str = "manual"
    confidence: float | None = None
    iteration: int = 0


class AnnotationBatch(BaseModel):
    annotations: list[AnnotationItem]


@app.post("/api/annotations")
async def api_save_annotations(req: AnnotationBatch):
    """保存/批量保存标注"""
    try:
        items = [a.model_dump(exclude_none=False) for a in req.annotations]
        results = save_batch(items)
        return {"status": "ok", "count": len(results), "annotations": results}
    except Exception as e:
        logger.exception(f"Save annotations error: {e}")
        raise HTTPException(status_code=500, detail=f"保存标注失败: {str(e)}")


@app.get("/api/annotations")
async def api_load_annotations(session_id: str):
    """加载指定 session 的所有标注"""
    try:
        annotations = load_by_session(session_id)
        return {"status": "ok", "session_id": session_id, "count": len(annotations), "annotations": annotations}
    except Exception as e:
        logger.exception(f"Load annotations error: {e}")
        raise HTTPException(status_code=500, detail=f"加载标注失败: {str(e)}")


@app.delete("/api/annotations/{annot_id}")
async def api_delete_annotation(annot_id: str):
    """删除单条标注"""
    try:
        ok = annotation_delete(annot_id)
        if not ok:
            raise HTTPException(status_code=404, detail="标注不存在")
        return {"status": "ok", "deleted": annot_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Delete annotation error: {e}")
        raise HTTPException(status_code=500, detail=f"删除标注失败: {str(e)}")


@app.get("/api/annotations/export")
async def api_export_annotations(session_id: str, format: str = "geojson"):
    """导出标注数据（geojson 或 coco 格式）"""
    try:
        if format == "coco":
            result = annotation_export_coco(session_id)
        else:
            result = annotation_export_geojson(session_id)
        return result
    except Exception as e:
        logger.exception(f"Export annotations error: {e}")
        raise HTTPException(status_code=500, detail=f"导出标注失败: {str(e)}")


@app.get("/api/annotations/sessions")
async def api_list_annotation_sessions():
    """列出所有标注 session"""
    try:
        sessions = list_sessions()
        return {"status": "ok", "sessions": sessions}
    except Exception as e:
        logger.exception(f"List sessions error: {e}")
        raise HTTPException(status_code=500, detail=f"列出 session 失败: {str(e)}")


@app.delete("/api/annotations/sessions/{session_id}")
async def api_delete_annotation_session(session_id: str):
    """删除整个 session 的标注"""
    try:
        count = delete_by_session(session_id)
        return {"status": "ok", "session_id": session_id, "deleted": count}
    except Exception as e:
        logger.exception(f"Delete session error: {e}")
        raise HTTPException(status_code=500, detail=f"删除 session 失败: {str(e)}")


# ==============================
# 工作流会话 API
# ==============================

class WorkflowStartRequest(BaseModel):
    title: str | None = "ReSAM 标注会话"


@app.post("/api/workflow/start")
async def api_workflow_start(req: WorkflowStartRequest):
    """
    创建标注工作流会话，返回 session_id。
    复用现有 sessions.db 基础设施，会话类型标注为 "annotation"。
    """
    try:
        import uuid
        session_id = uuid.uuid4().hex[:12]

        # 在标注表创建空 session 占位
        from tools.annotation_store import save, _now_iso
        placeholder = {
            "id": f"_workflow_init_{session_id}",
            "session_id": session_id,
            "image_path": "",
            "label": "__workflow__",
            "class_id": None,
            "geometry": json.dumps({"type": "Point", "coordinates": [0, 0]}),
            "mask_path": None,
            "source": "system",
            "confidence": None,
            "iteration": 0,
            "created_at": _now_iso(),
        }
        save(placeholder)

        # 可选：在 sessions.db 中同步创建记录
        try:
            from tools.schema_manager import get_conn as get_schema_conn
            schema_conn = get_schema_conn()
            cur = schema_conn.cursor()
            cur.execute(
                "INSERT OR IGNORE INTO sessions (id, title, type, created_at, updated_at) VALUES (?, ?, ?, datetime('now'), datetime('now'))",
                (session_id, req.title or "ReSAM 标注会话", "annotation")
            )
            schema_conn.commit()
        except Exception:
            pass  # sessions 表可能不存在，忽略

        logger.info(f"[Workflow] 创建标注会话: {session_id}, title={req.title}")
        return {"status": "ok", "session_id": session_id, "title": req.title}
    except Exception as e:
        logger.exception(f"Workflow start error: {e}")
        raise HTTPException(status_code=500, detail=f"创建工作流会话失败: {str(e)}")


# ==============================
# SSA 训练任务 API
# ==============================

# 训练任务状态追踪（内存中，服务重启后丢失属正常行为）
_training_tasks: dict = {}


class SAMTrainRequest(BaseModel):
    session_ids: list[str] = []       # 标注 session IDs
    epochs: int = 20
    checkpoint_name: str = "buildings_v1"
    image_dir: str = ""               # 影像目录（可选，默认使用 LOCAL_TIF 裁剪区域）
    base_checkpoint: str = ""         # 基底模型路径（空 = 从零训练）
    lora_rank: int = 4                # LoRA 秩（ReSAM）
    training_method: str = "resam"    # 训练方法: resam | ssa


@app.post("/api/sam-train")
async def sam_train(req: SAMTrainRequest):
    """
    提交 SSA 微调训练任务（异步后台执行）。

    从标注数据库导出数据 → 裁剪对应区域影像 → 执行 ssa_train.py
    返回 task_id 用于进度查询。
    """
    import subprocess
    import uuid
    import threading
    import sys

    task_id = uuid.uuid4().hex[:12]
    _training_tasks[task_id] = {
        "status": "pending",
        "epoch": 0,
        "total_epochs": req.epochs,
        "loss": None,
        "val_loss": None,
        "progress_pct": 0,
        "message": "任务已创建，等待启动...",
        "output_path": "",
        "error": None,
    }

    def _run_training():
        """后台执行训练流程"""
        task = _training_tasks[task_id]
        task["status"] = "running"
        task["message"] = "正在准备训练数据..."

        try:
            # 1. 从标注数据库导出训练数据
            annotations_dir = tempfile.mkdtemp(prefix="ssa_train_data_")
            annotations_json = os.path.join(annotations_dir, "annotations.json")

            all_annotations = []
            for sid in req.session_ids:
                sess_anns = load_by_session(sid)
                # 转换为训练格式
                for ann in sess_anns:
                    all_annotations.append({
                        "image_path": ann.get("image_path", ""),
                        "label": ann.get("label", ""),
                        "geometry": ann.get("geometry", {}),
                        "mask_path": ann.get("mask_path", ""),
                    })

            if not all_annotations:
                task["status"] = "failed"
                task["error"] = "无标注数据可训练（session 为空或无有效标注）"
                task["message"] = "训练失败：无数据"
                return

            with open(annotations_json, 'w') as f:
                json.dump(all_annotations, f)

            task["message"] = f"已导出 {len(all_annotations)} 条标注，启动训练..."
            logger.info(f"[SAM-Train] task={task_id}, 导出 {len(all_annotations)} 条标注")

            # 2. 确定影像目录（使用本地 TIF 或请求指定目录）
            image_dir = req.image_dir or "/home/server/python/GIS/output"
            if not os.path.isdir(image_dir):
                image_dir = tempfile.mkdtemp(prefix="ssa_images_")

            # 3. 执行训练脚本
            # 持久化存储 checkpoint
            ckpt_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "checkpoints")
            os.makedirs(ckpt_dir, exist_ok=True)
            output_path = os.path.join(ckpt_dir, f"sam3_ssa_{req.checkpoint_name}.pt")
            task["output_path"] = output_path

            python_bin = sys.executable

            # 根据训练方法选择脚本
            if req.training_method == "resam":
                train_script = "/home/server/python/map_assistant_v1/backend/tools/resam_train.py"
                cmd = [
                    python_bin, "-u", train_script,
                    "--annotations", annotations_json,
                    "--image-dir", image_dir,
                    "--epochs", str(req.epochs),
                    "--batch-size", "2",
                    "--lora-rank", str(req.lora_rank),
                    "--output", output_path,
                ]
            else:
                train_script = "/home/server/python/map_assistant_v1/backend/tools/ssa_train.py"
                cmd = [
                    python_bin, "-u", train_script,
                    "--annotations", annotations_json,
                    "--image-dir", image_dir,
                    "--epochs", str(req.epochs),
                    "--batch-size", "4",
                    "--output", output_path,
                ]

            # 基底模型（在已有 checkpoint 基础上继续训练）
            if req.base_checkpoint and os.path.exists(req.base_checkpoint):
                cmd.extend(["--base-checkpoint" if req.training_method != "resam" else "--resume", req.base_checkpoint])
                task["message"] = f"基于 {os.path.basename(req.base_checkpoint)} 继续训练..."

            logger.info(f"[SAM-Train] cmd: {' '.join(cmd)}")
            task["message"] = "训练中..."

            # 使用 Popen 实时读取输出，捕获 epoch 进度
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                env={**os.environ, "CUDA_VISIBLE_DEVICES": os.environ.get("CUDA_VISIBLE_DEVICES", "0")},
            )

            for line in proc.stdout:
                line = line.strip()
                # 解析 epoch 进度: "[ReSAM] Epoch 02/20 ..." 或 "[SSA-Train] Epoch 02/10 ..."
                if "Epoch" in line and "/" in line:
                    try:
                        parts = line.split("|")
                        ep_part = parts[0].strip()  # "Epoch 02/10"
                        ep_str = ep_part.split()[1]  # "02/10"
                        cur_ep, total_ep = ep_str.split("/")
                        task["epoch"] = int(cur_ep)
                        task["total_epochs"] = int(total_ep)
                        task["progress_pct"] = int(int(cur_ep) / int(total_ep) * 100)

                        for part in parts[1:]:
                            part = part.strip()
                            if "train_loss=" in part:
                                task["loss"] = float(part.split("=")[1].split()[0])
                            elif "val_loss=" in part:
                                task["val_loss"] = float(part.split("=")[1].split()[0])
                    except Exception:
                        pass

                logger.info(f"[SAM-Train {task_id}] {line}")

            proc.wait()
            if proc.returncode != 0:
                task["status"] = "failed"
                task["error"] = f"训练脚本异常退出 (code={proc.returncode})"
                task["message"] = "训练失败"
            else:
                task["status"] = "complete"
                task["progress_pct"] = 100
                val_info = ""
                if task.get("val_loss") is not None:
                    val_info = f", val_loss={task['val_loss']:.4f}"
                task["message"] = f"训练完成！train_loss={task.get('loss', '—')}{val_info} | {output_path}"
                # 记录元数据（类别、时间、loss、标注数、验证信息）
                meta_path = output_path.replace(".pt", ".meta.json")
                try:
                    unique_labels = sorted(set(a.get("label", "") for a in all_annotations if a.get("label")))
                    meta = {
                        "name": req.checkpoint_name,
                        "classes": unique_labels,
                        "epochs": req.epochs,
                        "final_loss": task.get("loss"),
                        "val_loss": task.get("val_loss"),
                        "annotation_count": len(all_annotations),
                        "session_ids": req.session_ids,
                        "created_at": dt.now().isoformat(),
                    }
                    with open(meta_path, 'w') as f:
                        json.dump(meta, f, ensure_ascii=False)
                except Exception as me:
                    logger.warning(f"[SAM-Train] 元数据保存失败: {me}")
                logger.info(f"[SAM-Train] task={task_id} 完成: {output_path}")

        except Exception as e:
            logger.exception(f"[SAM-Train] task={task_id} 异常: {e}")
            task["status"] = "failed"
            task["error"] = str(e)
            task["message"] = f"训练异常: {str(e)[:100]}"

    # 启动后台线程
    thread = threading.Thread(target=_run_training, daemon=True)
    thread.start()

    return {"task_id": task_id, "status": "pending"}


@app.get("/api/sam-train/{task_id}/status")
async def sam_train_status(task_id: str):
    """查询 SSA 训练任务进度"""
    task = _training_tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="训练任务不存在或已过期")
    return {
        "task_id": task_id,
        **task,
    }


# ==============================
# SSA Checkpoint 版本管理 API
# ==============================

CHECKPOINT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "checkpoints")
ACTIVE_CONFIG = os.path.join(CHECKPOINT_DIR, "active.json")


def _get_active_checkpoint():
    """读取当前激活的 checkpoint 路径"""
    try:
        if os.path.exists(ACTIVE_CONFIG):
            with open(ACTIVE_CONFIG, 'r') as f:
                data = json.load(f)
            active_path = data.get("checkpoint", "")
            if active_path and os.path.exists(active_path):
                return data
    except Exception:
        pass
    return {"checkpoint": "", "name": "默认 (SAM3 原生)"}


def _set_active_checkpoint(checkpoint_path, name):
    """写入激活的 checkpoint 配置"""
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)
    data = {"checkpoint": checkpoint_path, "name": name, "set_at": dt.now().isoformat()}
    with open(ACTIVE_CONFIG, 'w') as f:
        json.dump(data, f, ensure_ascii=False)
    # 同步更新环境变量，让 sam_predict 模块感知
    os.environ["SAM3_SSA_CHECKPOINT"] = checkpoint_path
    return data


# ==============================
# 点击提示分割 API (ReSAM)
# ==============================

class PointPredictRequest(BaseModel):
    points: list = []           # [[x, y], ...] 像素坐标
    labels: list = []           # [1, 0, ...] 1=正提示，0=负提示
    image_bounds: dict = {}     # {"north", "south", "east", "west"}
    image_path: str = ""        # 可选：直接指定图片路径
    session_id: str = ""        # 当前标注 session
    prompt: str = "object"      # 文本提示词


@app.post("/api/sam-predict-point")
async def sam_predict_point(req: PointPredictRequest):
    """
    基于点提示的 SAM 分割。
    用户在地图上点击正点/负点，根据点位置生成分割 mask。

    通过子进程调用 sam 环境运行 sam_predict.py --point，
    返回 GeoJSON 多边形（与文字提示接口格式一致）。
    """
    if not req.points or not req.labels:
        return {"polygons": [], "message": "无提示点"}

    if len(req.points) != len(req.labels):
        raise HTTPException(status_code=400, detail="points 和 labels 数量不匹配")

    # 如果前端未提供 image_bounds，从点坐标自动计算（带约 300m 边距）
    if not req.image_bounds:
        xs = [p[0] for p in req.points]
        ys = [p[1] for p in req.points]
        margin = 0.003  # ~300m
        req.image_bounds = {
            "west": min(xs) - margin,
            "east": max(xs) + margin,
            "south": min(ys) - margin,
            "north": max(ys) + margin,
        }

    try:
        import subprocess as _sp

        sam_script = "/home/server/python/map_assistant_v1/backend/tools/sam_predict.py"
        python_bin = "/home/server/miniconda3/envs/sam/bin/python"

        # 构建输入 JSON
        input_data = {
            "points": req.points,
            "labels": req.labels,
            "image_bounds": req.image_bounds if req.image_bounds else None,
            "image_path": req.image_path if req.image_path else None,
            "prompt": req.prompt or "object",
        }

        cmd = [python_bin, "-u", sam_script, "--point"]
        logger.info(f"[SAM-Point] cmd: {' '.join(cmd)}, points={len(req.points)}")

        proc = _sp.run(
            cmd,
            input=json.dumps(input_data),
            capture_output=True,
            text=True,
            timeout=60,
            env={**os.environ, "CUDA_VISIBLE_DEVICES": os.environ.get("CUDA_VISIBLE_DEVICES", "0")},
        )

        if proc.returncode != 0:
            err_msg = (proc.stderr or "")[:500]
            logger.error(f"[SAM-Point] 子进程失败: {err_msg}")
            raise HTTPException(status_code=500, detail=f"点提示分割失败: {err_msg[:200]}")

        # 解析 stdout 中的最后一行 JSON
        stdout_lines = [l for l in proc.stdout.strip().splitlines() if l.strip()]
        if not stdout_lines:
            return {"polygons": [], "message": "模型无输出"}

        result = json.loads(stdout_lines[-1])
        return result

    except _sp.TimeoutExpired:
        logger.error("[SAM-Point] 子进程超时")
        raise HTTPException(status_code=504, detail="点提示分割超时")
    except json.JSONDecodeError as e:
        logger.error(f"[SAM-Point] 解析输出失败: {e}")
        raise HTTPException(status_code=500, detail="点提示分割输出解析失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[SAM-Point] 点提示分割失败: {e}")
        raise HTTPException(status_code=500, detail=f"点提示分割失败: {str(e)[:200]}")


@app.get("/api/sam-checkpoints")
async def list_checkpoints():
    """列出所有 SSA checkpoint 版本"""
    items = []
    # 默认选项
    items.append({
        "name": "默认 (SAM3 原生)",
        "path": "",
        "classes": [],
        "epochs": 0,
        "annotation_count": 0,
        "created_at": "",
        "active": not bool(_get_active_checkpoint().get("checkpoint")),
    })
    active_info = _get_active_checkpoint()
    active_path = active_info.get("checkpoint", "")

    try:
        os.makedirs(CHECKPOINT_DIR, exist_ok=True)
        for fname in sorted(os.listdir(CHECKPOINT_DIR), reverse=True):
            if not fname.endswith(".pt"):
                continue
            ckpt_path = os.path.join(CHECKPOINT_DIR, fname)
            meta_path = ckpt_path.replace(".pt", ".meta.json")
            name = fname.replace("sam3_ssa_", "").replace(".pt", "")

            # 读取元数据
            meta = {"classes": [], "epochs": 0, "annotation_count": 0, "created_at": ""}
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, 'r') as f:
                        meta = json.load(f)
                except Exception:
                    pass

            chart_file = ckpt_path.replace(".pt", "_chart.png")
            # 检测训练方法
            training_method = meta.get("training_method", "ssa")
            # 也可以从 checkpoint 文件本身检测
            try:
                import torch as _t
                _ckpt = _t.load(ckpt_path, map_location='cpu')
                if 'lora' in _ckpt and 'training_method' in _ckpt:
                    training_method = 'ReSAM'
                    meta['lora_rank'] = _ckpt.get('metadata', {}).get('lora_rank', 4)
            except Exception:
                pass

            items.append({
                "name": meta.get("name", name),
                "path": ckpt_path,
                "classes": meta.get("classes", []),
                "epochs": meta.get("epochs", 0),
                "annotation_count": meta.get("annotation_count", 0),
                "final_loss": meta.get("final_loss"),
                "val_loss": meta.get("val_loss"),
                "created_at": meta.get("created_at", ""),
                "active": ckpt_path == active_path,
                "has_chart": os.path.exists(chart_file),
                "chart_url": f"/api/sam-checkpoints/{name}/chart" if os.path.exists(chart_file) else None,
                "training_method": training_method,
                "lora_rank": meta.get("lora_rank"),
            })
    except Exception as e:
        logger.warning(f"列出 checkpoint 失败: {e}")

    return {"checkpoints": items, "active": active_info}


class CheckpointActivateRequest(BaseModel):
    path: str = ""   # 空字符串 = 恢复默认（不使用 adapter）


@app.post("/api/sam-checkpoints/activate")
async def activate_checkpoint(req: CheckpointActivateRequest):
    """切换激活的 SSA checkpoint 版本"""
    if req.path and not os.path.exists(req.path):
        raise HTTPException(status_code=404, detail=f"checkpoint 不存在: {req.path}")

    if not req.path:
        # 恢复默认
        _set_active_checkpoint("", "默认 (SAM3 原生)")
        # 清除模型缓存，让下次推理使用原生 SAM3
        try:
            from tools.sam_predict import reload_sam_model
            reload_sam_model()
        except Exception:
            pass
        logger.info("[Checkpoint] 已恢复为默认 SAM3 原生模式")
        return {"status": "ok", "active": {"checkpoint": "", "name": "默认 (SAM3 原生)"}}

    name = os.path.basename(req.path).replace("sam3_ssa_", "").replace(".pt", "")
    result = _set_active_checkpoint(req.path, name)
    # 清除模型缓存，让下次推理使用新 checkpoint
    try:
        from tools.sam_predict import reload_sam_model
        reload_sam_model()
    except Exception:
        pass
    logger.info(f"[Checkpoint] 已激活: {name} ({req.path})")
    return {"status": "ok", "active": result}


@app.get("/api/sam-checkpoints/active")
async def get_active_checkpoint():
    """查询当前激活的 checkpoint"""
    return _get_active_checkpoint()


class CheckpointRenameRequest(BaseModel):
    path: str
    new_name: str


@app.post("/api/sam-checkpoints/rename")
async def rename_checkpoint(req: CheckpointRenameRequest):
    """重命名 checkpoint 的显示名称（写入 meta.json）"""
    if not req.path or not os.path.exists(req.path):
        raise HTTPException(status_code=404, detail="checkpoint 不存在")
    meta_path = req.path.replace(".pt", ".meta.json")
    meta = {}
    if os.path.exists(meta_path):
        try:
            with open(meta_path, 'r') as f:
                meta = json.load(f)
        except Exception:
            pass
    meta["name"] = req.new_name.strip()
    with open(meta_path, 'w') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    logger.info(f"[Checkpoint] 重命名: {req.path} -> {req.new_name}")
    return {"status": "ok", "name": meta["name"]}


@app.delete("/api/sam-checkpoints")
async def delete_checkpoint(path: str):
    """删除指定的 checkpoint 及其关联文件"""
    if not path:
        raise HTTPException(status_code=400, detail="路径不能为空")
    # 安全检查：必须在 checkpoint 目录内
    if not path.startswith(CHECKPOINT_DIR):
        raise HTTPException(status_code=403, detail="不允许删除该路径")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 不允许删除当前激活的 checkpoint
    active_info = _get_active_checkpoint()
    if active_info.get("checkpoint") == path:
        raise HTTPException(status_code=400, detail="不能删除当前激活的模型，请先切换到其他模型")
    
    # 删除关联文件: .pt, .meta.json, _chart.png, _best.pt, _best_chart.png
    base = path.replace(".pt", "")
    deleted = []
    for suffix in [".pt", ".meta.json", "_chart.png", "_best.pt", "_best_chart.png", "_best.meta.json"]:
        fp = base + suffix
        if os.path.exists(fp):
            os.remove(fp)
            deleted.append(os.path.basename(fp))
    
    logger.info(f"[Checkpoint] 删除: {deleted}")
    return {"status": "ok", "deleted": deleted}


@app.get("/api/sam-checkpoints/{checkpoint_name}/chart")
async def get_checkpoint_chart(checkpoint_name: str):
    """返回指定 checkpoint 的训练曲线图表（PNG）"""
    from fastapi.responses import FileResponse
    
    # 安全检查：防止路径遍历
    safe_name = os.path.basename(checkpoint_name)
    chart_path = os.path.join(CHECKPOINT_DIR, f"sam3_ssa_{safe_name}_chart.png")
    
    if not os.path.exists(chart_path):
        # 尝试不带 sam3_ssa_ 前缀的路径
        alt_path = os.path.join(CHECKPOINT_DIR, f"{safe_name}_chart.png")
        if os.path.exists(alt_path):
            chart_path = alt_path
        else:
            raise HTTPException(status_code=404, detail=f"图表不存在: {safe_name}")
    
    return FileResponse(chart_path, media_type="image/png",
                         headers={"Cache-Control": "no-cache"})


if __name__ == "__main__":
    import uvicorn
    import socket
    _port = int(os.environ.get("PORT") or os.environ.get("APP_PORT") or "8006")
    # 创建带 SO_REUSEADDR 的 socket，避免 PM2 重启时端口抢占导致启动失败
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    sock.bind(("0.0.0.0", _port))
    sock.listen(2048)
    uvicorn.run(app, fd=sock.fileno())
